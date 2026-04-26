from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet
from flask import Flask, render_template, jsonify, request, send_file, redirect
import pandas as pd
import sqlite3
import time
from io import BytesIO
from collections import Counter
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
import os
import json
from datetime import datetime
from zoneinfo import ZoneInfo
from werkzeug.utils import secure_filename
from calendar import month_name
from auth_blueprint import configure_auth_app, get_current_user, usuario_pode
from flask import abort

app = Flask(__name__)
app.secret_key = os.getenv("SECRET_KEY", "troque-esta-chave-em-producao")
app.config["AUTH_DB_PATH"] = os.path.join(app.instance_path, "auth_users.db")
os.makedirs(app.instance_path, exist_ok=True)
os.environ.setdefault("ADMIN_EMAIL", "Vitordf57")
os.environ.setdefault("ADMIN_PASSWORD", "Condornew25@")
configure_auth_app(app)

SCREEN_ENDPOINT_RULES = {
    "/home": "home",
    "/metricas-full": "metricas_full",
    "/dados-fiscais": "metricas_full",
    "/conferencia": "conferencia",
    "/picking": "picking",
    "/gestao-operacional": "picking",
    "/embalagem": "conferencia",
    "/api/embalagem/confirmar": "conferencia",
    "/api/historico-mensal": "principal",
    "/api/historico-mensal-resumo": "principal",
    "/dados": "principal",
    "/dados-dashboard": "principal",
    "/exportar-excel": "principal",
    "/criar-lote-enviando": "principal",
    "/gerar-filete": "principal",
    "/gerar-filete-enviando": "principal",
    "/dados-full": "principal",
    "/salvar-status": "principal",
    "/status": "principal",
    "/comentarios": "principal",
    "/comentarios-mlb": "principal",
    "/comentarios-mlb-chat": "principal",
    "/salvar-comentario": "principal",
    "/salvar-comentario-mlb": "principal",
    "/debug-comentarios": "principal",
    "/dashboard": "principal",
    "/gerar-pdf-filete": "principal",
    "/api/full-distribuicao": "principal",
    "/api/full-distribuicao-detalhe": "principal",
    "/salvar-conferencia-item": "conferencia",
    "/api/picking/coletar": "picking",
}

SCREEN_PREFIX_RULES = [
    ("/api/coletas-calendario", "metricas_full"),
    ("/api/lote-envio/", "metricas_full"),
    ("/lote-envio/", "metricas_full"),
    ("/excluir-lote/", "metricas_full"),
    ("/embalagem/", "conferencia"),
]


def _resolver_tela_por_path(path: str):
    if path == "/":
        tela = request.args.get("tela", "principal") or "principal"
        return tela

    if path in SCREEN_ENDPOINT_RULES:
        return SCREEN_ENDPOINT_RULES[path]

    for prefix, screen_key in SCREEN_PREFIX_RULES:
        if path.startswith(prefix):
            return screen_key

    return None


@app.before_request
def proteger_rotas_com_login_e_permissao():
    endpoint = request.endpoint or ""
    path = request.path or "/"

    if endpoint.startswith("auth_bp.") or endpoint == "static":
        return None

    tela = _resolver_tela_por_path(path)
    if tela is None:
        return None

    user = get_current_user()
    if not user:
        return redirect(f"/login?next={request.full_path.rstrip('?')}")

    if not (user.is_admin or usuario_pode(tela)):
        abort(403)

    return None


UPLOAD_FOLDER_CONFERENCIA = os.path.join("static", "uploads", "conferencia")
os.makedirs(UPLOAD_FOLDER_CONFERENCIA, exist_ok=True)

app.config["UPLOAD_FOLDER_CONFERENCIA"] = UPLOAD_FOLDER_CONFERENCIA

CSV_URL = "https://docs.google.com/spreadsheets/d/1DKdRHI9IEacgOwsEd-bnAN4nU3dA_clULxU1mFa8LmY/export?format=csv&gid=0"
CSV_URL_FULL = "https://docs.google.com/spreadsheets/d/1DKdRHI9IEacgOwsEd-bnAN4nU3dA_clULxU1mFa8LmY/export?format=csv&gid=184771586"

CACHE_TTL = 300

cache_dados = None
cache_dados_ts = 0

cache_full = None
cache_full_ts = 0


def carregar_dados_base():
    global cache_dados, cache_dados_ts

    agora = time.time()

    if cache_dados is None or (agora - cache_dados_ts) > CACHE_TTL:
        df = pd.read_csv(CSV_URL)
        df = df.fillna("")
        cache_dados = df
        cache_dados_ts = agora

    return cache_dados.copy()


def montar_mapa_base_por_codigo_sku(df_base):
    mapa_codigo = {}
    mapa_sku = {}

    if df_base is None or df_base.empty:
        return mapa_codigo, mapa_sku

    df_base = df_base.fillna("").copy()

    for _, row in df_base.iterrows():
        dados = {str(col): row[col] for col in df_base.columns}

        codigo = str(row.get("Código do Anúncio", "") or "").strip()
        sku = str(row.get("SKU", "") or "").strip()

        if codigo and codigo not in mapa_codigo:
            mapa_codigo[codigo] = dados

        if sku and sku not in mapa_sku:
            mapa_sku[sku] = dados

    return mapa_codigo, mapa_sku


def enriquecer_itens_lote_com_base(itens):
    if not itens:
        return itens

    try:
        df_base = carregar_dados_base()
    except:
        return itens

    mapa_codigo, mapa_sku = montar_mapa_base_por_codigo_sku(df_base)

    itens_enriquecidos = []
    for item in itens:
        item_dict = dict(item)

        codigo = str(item_dict.get("codigo", "") or "").strip()
        sku = str(item_dict.get("sku", "") or "").strip()

        base_item = mapa_codigo.get(codigo) or mapa_sku.get(sku) or {}

        dados_json_atual = {}
        try:
            dados_json_atual = json.loads(item_dict.get("dados_json") or "{}")
            if not isinstance(dados_json_atual, dict):
                dados_json_atual = {}
        except:
            dados_json_atual = {}

        dados_mesclados = {}
        if base_item:
            for k, v in base_item.items():
                dados_mesclados[str(k)] = v
        for k, v in dados_json_atual.items():
            if v not in [None, "", []]:
                dados_mesclados[str(k)] = v

        if codigo:
            dados_mesclados["Código do Anúncio"] = codigo
        if sku:
            dados_mesclados["SKU"] = sku
        if item_dict.get("titulo"):
            dados_mesclados["Título"] = item_dict.get("titulo")
        elif dados_mesclados.get("Título"):
            item_dict["titulo"] = dados_mesclados.get("Título")

        if item_dict.get("nickname"):
            dados_mesclados["Nickname"] = item_dict.get("nickname")
        elif dados_mesclados.get("Nickname"):
            item_dict["nickname"] = dados_mesclados.get("Nickname")

        if item_dict.get("endereco"):
            dados_mesclados["ENDEREÇO"] = item_dict.get("endereco")
        elif dados_mesclados.get("ENDEREÇO"):
            item_dict["endereco"] = dados_mesclados.get("ENDEREÇO")

        if item_dict.get("lote_filete"):
            dados_mesclados["LOTE"] = item_dict.get("lote_filete")
        elif dados_mesclados.get("LOTE"):
            item_dict["lote_filete"] = dados_mesclados.get("LOTE")

        item_dict["dados_json"] = json.dumps(dados_mesclados, ensure_ascii=False)
        itens_enriquecidos.append(item_dict)

    return itens_enriquecidos


def carregar_csv_com_cache(url, tipo="dados"):
    global cache_dados, cache_dados_ts, cache_full, cache_full_ts

    agora = time.time()

    if tipo == "dados":
        if cache_dados is None or (agora - cache_dados_ts) > CACHE_TTL:
            df = pd.read_csv(url)
            df = df.fillna("")
            cache_dados = df.to_dict(orient="records")
            cache_dados_ts = agora
        return cache_dados

    if tipo == "full":
        if cache_full is None or (agora - cache_full_ts) > CACHE_TTL:
            df = pd.read_csv(url)
            df = df.fillna("")
            cache_full = df.to_dict(orient="records")
            cache_full_ts = agora
        return cache_full

    return []


def normalizar_texto(valor):
    return str(valor or "").strip().upper()


def numero_float(valor):
    if valor is None or valor == "":
        return 0
    try:
        return float(str(valor).replace(".", "").replace(",", "."))
    except:
        return 0


TIMELINE_ETAPAS = [
    "MLBS EM ANÁLISE",
    "CRIAR PEDIDO SIGNUS",
    "GERAR LOTE MELI",
    "AGENDAR DATA",
    "CONFERIR QUANTIDADES",
    "EM SEPARAÇÃO",
    "EM CONFERÊNCIA",
    "EMBALAGEM",
    "CONFERIR CAIXAS MASTER",
    "AGUARDANDO COLETA",
    "COLETADO"
]


ETAPA_TO_STATUS_FIELD = {
    "MLBS EM ANÁLISE": "status_ecommerce",
    "CRIAR PEDIDO SIGNUS": "status_ecommerce",
    "GERAR LOTE MELI": "status_ecommerce",
    "AGENDAR DATA": "status_ecommerce",
    "CONFERIR QUANTIDADES": "status_ecommerce",
    "EM SEPARAÇÃO": "status_expedicao",
    "EM CONFERÊNCIA": "status_expedicao",
    "EMBALAGEM": "status_expedicao",
    "CONFERIR CAIXAS MASTER": "status_expedicao",
    "AGUARDANDO COLETA": "status_expedicao",
    "COLETADO": "status_expedicao"
}


def agora_str():
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def parse_data_hora(valor):
    texto = str(valor or "").strip()
    if not texto:
        return None

    formatos = [
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d %H:%M",
        "%Y-%m-%d",
        "%d/%m/%Y %H:%M:%S",
        "%d/%m/%Y %H:%M",
        "%d/%m/%Y"
    ]

    for formato in formatos:
        try:
            return datetime.strptime(texto, formato)
        except:
            pass

    return None


def formatar_data_hora_br(valor, vazio="-"):
    dt = parse_data_hora(valor)
    if not dt:
        return vazio
    return dt.strftime("%d/%m/%Y %H:%M")


def formatar_data_br(valor, vazio="-"):
    dt = parse_data_hora(valor)
    if not dt:
        return vazio
    return dt.strftime("%d/%m/%Y")


def agora_brasilia():
    try:
        return datetime.now(ZoneInfo("America/Sao_Paulo"))
    except:
        return datetime.now()


def agora_str_brasilia():
    return agora_brasilia().strftime("%Y-%m-%d %H:%M:%S")


def montar_historico_comentarios_mlb(cursor, codigo):
    codigo = str(codigo or "").strip()
    if not codigo:
        return []

    try:
        cursor.execute("""
            SELECT id, mensagem, data_hora
            FROM comentarios_mlb_chat
            WHERE codigo = ?
            ORDER BY COALESCE(data_hora, ''), id
        """, (codigo,))
        rows = cursor.fetchall()
    except:
        rows = []

    historico = []
    for row in rows:
        if isinstance(row, sqlite3.Row):
            mensagem = row["mensagem"]
            data_hora = row["data_hora"]
        else:
            mensagem = row[1] if len(row) > 1 else ""
            data_hora = row[2] if len(row) > 2 else ""

        mensagem = str(mensagem or "").strip()
        data_hora = str(data_hora or "").strip()
        if not mensagem:
            continue

        historico.append({
            "mensagem": mensagem,
            "data_hora": data_hora,
            "data_hora_br": formatar_data_hora_br(data_hora, "")
        })

    if historico:
        return historico

    try:
        cursor.execute("SELECT comentario FROM comentarios_mlb WHERE codigo = ?", (codigo,))
        row_legado = cursor.fetchone()
    except:
        row_legado = None

    comentario_legado = ""
    if row_legado is not None:
        if isinstance(row_legado, sqlite3.Row):
            comentario_legado = str(row_legado["comentario"] or "").strip()
        elif isinstance(row_legado, (list, tuple)) and len(row_legado) > 0:
            comentario_legado = str(row_legado[0] or "").strip()
        else:
            comentario_legado = str(row_legado or "").strip()

    if comentario_legado:
        data_hora_legado = agora_str_brasilia()
        try:
            cursor.execute("""
                INSERT INTO comentarios_mlb_chat (codigo, mensagem, data_hora)
                VALUES (?, ?, ?)
            """, (codigo, comentario_legado, data_hora_legado))
        except:
            pass

        historico.append({
            "mensagem": comentario_legado,
            "data_hora": data_hora_legado,
            "data_hora_br": formatar_data_hora_br(data_hora_legado, "")
        })

    return historico


def timeline_vazio():
    return {etapa: "" for etapa in TIMELINE_ETAPAS}


def carregar_timeline_json(valor):
    timeline = timeline_vazio()

    if not valor:
        return timeline

    try:
        dados = json.loads(valor)
        if isinstance(dados, dict):
            for etapa in TIMELINE_ETAPAS:
                timeline[etapa] = str(dados.get(etapa, "") or "").strip()
    except:
        pass

    return timeline


def timeline_para_json(timeline):
    dados = {etapa: str(timeline.get(etapa, "") or "").strip() for etapa in TIMELINE_ETAPAS}
    return json.dumps(dados, ensure_ascii=False)


def indice_etapa(etapa):
    try:
        return TIMELINE_ETAPAS.index(etapa)
    except ValueError:
        return 0


def sincronizar_timeline_ate_etapa(timeline, etapa_atual, timestamp=None):
    if timestamp is None:
        timestamp = agora_str()

    idx = indice_etapa(etapa_atual)
    for i, etapa in enumerate(TIMELINE_ETAPAS):
        if i <= idx and not str(timeline.get(etapa, "") or "").strip():
            timeline[etapa] = timestamp
    return timeline


def obter_status_abertura_por_etapa(etapa_atual):
    if indice_etapa(etapa_atual) <= indice_etapa("CONFERIR QUANTIDADES"):
        return "ABERTO"
    return "FECHADO"


def calcular_lead_time_segundos(timeline):
    momentos = []
    for etapa in TIMELINE_ETAPAS:
        dt = parse_data_hora(timeline.get(etapa, ""))
        if dt:
            momentos.append(dt)

    if not momentos:
        return 0

    inicio = min(momentos)
    fim = max(momentos)
    return max(0, int((fim - inicio).total_seconds()))


def formatar_duracao_humana(segundos):
    segundos = int(segundos or 0)
    if segundos <= 0:
        return "-"

    dias = segundos // 86400
    horas = (segundos % 86400) // 3600
    minutos = (segundos % 3600) // 60

    partes = []
    if dias:
        partes.append(f"{dias}d")
    if horas:
        partes.append(f"{horas}h")
    if minutos or not partes:
        partes.append(f"{minutos}min")
    return " ".join(partes)


def resumir_contas_lote(itens_lote):
    mapa = {}
    for item in itens_lote:
        conta = str(item["nickname"] or "SEM CONTA").strip() or "SEM CONTA"
        if conta not in mapa:
            mapa[conta] = {"mlbs": set(), "pecas": 0}
        mapa[conta]["mlbs"].add(str(item["codigo"] or ""))
        try:
            mapa[conta]["pecas"] += int(item["quantidade"] or 0)
        except:
            pass

    linhas = []
    for conta, dados in sorted(mapa.items(), key=lambda x: x[0]):
        linhas.append({
            "conta": conta,
            "mlbs": len(dados["mlbs"]),
            "pecas": int(dados["pecas"])
        })
    return linhas


def construir_timeline_exibicao(timeline, etapa_atual):
    idx_atual = indice_etapa(etapa_atual)
    etapas = []

    for idx, etapa in enumerate(TIMELINE_ETAPAS):
        timestamp = str(timeline.get(etapa, "") or "").strip()
        if timestamp:
            estado = "concluida"
        elif idx == idx_atual:
            estado = "atual"
        else:
            estado = "pendente"

        if idx < idx_atual and not timestamp:
            estado = "concluida"

        etapas.append({
            "nome": etapa,
            "horario": formatar_data_hora_br(timestamp, "-"),
            "estado": estado
        })

    return etapas


def atualizar_statuss_por_etapa(cursor, numero_lote, etapa_atual):
    status = obter_status_abertura_por_etapa(etapa_atual)
    status_expedicao = "AGUARDANDO"
    status_ecommerce = "AGUARDANDO"

    campo = ETAPA_TO_STATUS_FIELD.get(etapa_atual)
    if campo == "status_expedicao":
        status_expedicao = etapa_atual
    else:
        status_ecommerce = etapa_atual

    cursor.execute(
        """
        UPDATE lotes_envio
        SET status = ?,
            status_expedicao = ?,
            status_ecommerce = ?
        WHERE numero_lote = ?
        """,
        (status, status_expedicao, status_ecommerce, numero_lote)
    )


def garantir_timeline_lote(cursor, numero_lote, etapa_atual=None):
    cursor.execute(
        "SELECT timeline_json, etapa_atual FROM lotes_envio WHERE numero_lote = ?",
        (numero_lote,)
    )
    row = cursor.fetchone()
    if not row:
        return timeline_vazio(), etapa_atual or TIMELINE_ETAPAS[0]

    timeline = carregar_timeline_json(row[0] if not isinstance(row, sqlite3.Row) else row["timeline_json"])
    etapa_salva = (row[1] if not isinstance(row, sqlite3.Row) else row["etapa_atual"]) or TIMELINE_ETAPAS[0]
    etapa_final = etapa_atual or etapa_salva
    return timeline, etapa_final


def atualizar_etapa_lote(numero_lote, etapa_atual, data_coleta_agendada=""):
    conn = sqlite3.connect("status.db")
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()

    timeline, _ = garantir_timeline_lote(cursor, numero_lote, etapa_atual)
    timestamp = agora_str()
    timeline = sincronizar_timeline_ate_etapa(timeline, etapa_atual, timestamp)

    cursor.execute(
        """
        UPDATE lotes_envio
        SET etapa_atual = ?,
            timeline_json = ?,
            data_coleta_agendada = COALESCE(NULLIF(?, ''), data_coleta_agendada)
        WHERE numero_lote = ?
        """,
        (etapa_atual, timeline_para_json(timeline), str(data_coleta_agendada or "").strip(), numero_lote)
    )
    atualizar_statuss_por_etapa(cursor, numero_lote, etapa_atual)

    conn.commit()
    conn.close()


def sincronizar_picking_itens(numero_lote, df_lote):
    conn = sqlite3.connect("status.db")
    cursor = conn.cursor()

    existentes = {}
    cursor.execute("SELECT codigo, coletado, coletado_em, observacao FROM lotes_picking_itens WHERE numero_lote = ?", (numero_lote,))
    for codigo, coletado, coletado_em, observacao in cursor.fetchall():
        existentes[str(codigo)] = {
            "coletado": int(coletado or 0),
            "coletado_em": coletado_em or "",
            "observacao": observacao or ""
        }

    for _, item in df_lote.iterrows():
        codigo = str(item.get("Código do Anúncio", "") or item.get("codigo", "") or "")
        sku = str(item.get("SKU", "") or item.get("sku", "") or "")
        endereco = str(item.get("ENDEREÇO", "") or item.get("endereco", "") or "")
        titulo = str(item.get("Título", "") or item.get("titulo", "") or "")
        conta = str(item.get("Nickname", "") or item.get("nickname", "") or "")
        selo = str(item.get("SELO", "") or item.get("selo", "") or "")
        quantidade_base = item.get("Enviar", item.get("quantidade", 0))
        try:
            quantidade = int(float(quantidade_base or 0))
        except:
            quantidade = 0

        coletado = existentes.get(codigo, {}).get("coletado", 0)
        coletado_em = existentes.get(codigo, {}).get("coletado_em", "")
        observacao = existentes.get(codigo, {}).get("observacao", "")

        cursor.execute(
            """
            INSERT INTO lotes_picking_itens (
                numero_lote, codigo, sku, endereco, titulo, conta, selo, quantidade, observacao, coletado, coletado_em
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ON CONFLICT(numero_lote, codigo) DO UPDATE SET
                sku = excluded.sku,
                endereco = excluded.endereco,
                titulo = excluded.titulo,
                conta = excluded.conta,
                selo = excluded.selo,
                quantidade = excluded.quantidade,
                observacao = ?,
                coletado = ?,
                coletado_em = ?
            """,
            (numero_lote, codigo, sku, endereco, titulo, conta, selo, quantidade, observacao, coletado, coletado_em, observacao, coletado, coletado_em)
        )

    conn.commit()
    conn.close()

@app.route("/api/historico-mensal")
def api_historico_mensal():
    mlb = str(request.args.get("mlb", "")).strip().upper()
    mes = request.args.get("mes", "").strip()
    ano = request.args.get("ano", "").strip()

    if not mlb or not mes or not ano:
        return jsonify({
            "ok": False,
            "erro": "Parâmetros obrigatórios: mlb, mes e ano."
        }), 400

    try:
        mes = int(mes)
        ano = int(ano)
    except:
        return jsonify({
            "ok": False,
            "erro": "Mês e ano inválidos."
        }), 400

    csv_url_acompanhamento = "https://docs.google.com/spreadsheets/d/1DKdRHI9IEacgOwsEd-bnAN4nU3dA_clULxU1mFa8LmY/export?format=csv&gid=1492834688"

    try:
        df = pd.read_csv(csv_url_acompanhamento)
        df = df.fillna("")

        df.columns = [str(c).strip() for c in df.columns]

        coluna_mlb = "# de anúncio"
        coluna_data = "DATA DA VENDA"
        coluna_qtd = "UNIDADE VENDIDA"

        if coluna_mlb not in df.columns or coluna_data not in df.columns or coluna_qtd not in df.columns:
            return jsonify({
                "ok": False,
                "erro": "Colunas esperadas não encontradas na planilha."
            }), 400

        df[coluna_mlb] = df[coluna_mlb].astype(str).str.strip().str.upper()
        df[coluna_data] = pd.to_datetime(df[coluna_data], format="%d/%m/%Y", errors="coerce")
        df[coluna_qtd] = pd.to_numeric(df[coluna_qtd], errors="coerce").fillna(0)

        df_filtrado = df[
            (df[coluna_mlb] == mlb) &
            (df[coluna_data].dt.month == mes) &
            (df[coluna_data].dt.year == ano)
        ]

        total_unidades = int(df_filtrado[coluna_qtd].sum())
        total_pedidos = int(len(df_filtrado))

        return jsonify({
            "ok": True,
            "mlb": mlb,
            "mes": mes,
            "ano": ano,
            "mes_nome": month_name[mes],
            "unidades_vendidas": total_unidades,
            "pedidos": total_pedidos
        })

    except Exception as e:
        return jsonify({
            "ok": False,
            "erro": f"Erro ao processar histórico mensal: {str(e)}"
        }), 500

@app.route("/api/historico-mensal-resumo")
def api_historico_mensal_resumo():
    mlb = str(request.args.get("mlb", "")).strip().upper()

    if not mlb:
        return jsonify({"ok": False})

    url = "https://docs.google.com/spreadsheets/d/1DKdRHI9IEacgOwsEd-bnAN4nU3dA_clULxU1mFa8LmY/export?format=csv&gid=1492834688"

    df = pd.read_csv(url)
    df = df.fillna("")

    df["# de anúncio"] = df["# de anúncio"].astype(str).str.upper().str.strip()
    df["DATA DA VENDA"] = pd.to_datetime(df["DATA DA VENDA"], format="%d/%m/%Y", errors="coerce")
    df["UNIDADE VENDIDA"] = pd.to_numeric(df["UNIDADE VENDIDA"], errors="coerce").fillna(0)

    df = df[df["# de anúncio"] == mlb]

    df["MES"] = df["DATA DA VENDA"].dt.month

    resumo = df.groupby("MES")["UNIDADE VENDIDA"].sum().to_dict()

    nomes_meses = {
        1:"JANEIRO",2:"FEVEREIRO",3:"MARÇO",4:"ABRIL",
        5:"MAIO",6:"JUNHO",7:"JULHO",8:"AGOSTO",
        9:"SETEMBRO",10:"OUTUBRO",11:"NOVEMBRO",12:"DEZEMBRO"
    }

    resultado = []

    for i in range(1,13):
        resultado.append({
            "mes": nomes_meses[i],
            "valor": int(resumo.get(i, 0))
        })

    return jsonify({
        "ok": True,
        "dados": resultado
    })


def garantir_tabela_embalagem(cursor):
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS lotes_embalagem_itens (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            numero_lote TEXT,
            codigo TEXT,
            sku TEXT,
            titulo TEXT DEFAULT '',
            quantidade INTEGER DEFAULT 0,
            endereco TEXT DEFAULT '',
            lote_filete TEXT DEFAULT '',
            observacao TEXT DEFAULT '',
            embalado INTEGER DEFAULT 0,
            embalado_em TEXT DEFAULT '',
            embalado_por TEXT DEFAULT '',
            quantidade_embalada INTEGER DEFAULT 0,
            divergencia INTEGER DEFAULT 0,
            divergencia_em TEXT DEFAULT '',
            UNIQUE(numero_lote, codigo)
        )
    """)

    cursor.execute("PRAGMA table_info(lotes_embalagem_itens)")
    colunas = [col[1] for col in cursor.fetchall()]
    if "observacao" not in colunas:
        cursor.execute("ALTER TABLE lotes_embalagem_itens ADD COLUMN observacao TEXT DEFAULT ''")
    if "embalado" not in colunas:
        cursor.execute("ALTER TABLE lotes_embalagem_itens ADD COLUMN embalado INTEGER DEFAULT 0")
    if "embalado_em" not in colunas:
        cursor.execute("ALTER TABLE lotes_embalagem_itens ADD COLUMN embalado_em TEXT DEFAULT ''")
    if "embalado_por" not in colunas:
        cursor.execute("ALTER TABLE lotes_embalagem_itens ADD COLUMN embalado_por TEXT DEFAULT ''")
    if "quantidade_embalada" not in colunas:
        cursor.execute("ALTER TABLE lotes_embalagem_itens ADD COLUMN quantidade_embalada INTEGER DEFAULT 0")
    if "divergencia" not in colunas:
        cursor.execute("ALTER TABLE lotes_embalagem_itens ADD COLUMN divergencia INTEGER DEFAULT 0")
    if "divergencia_em" not in colunas:
        cursor.execute("ALTER TABLE lotes_embalagem_itens ADD COLUMN divergencia_em TEXT DEFAULT ''")



def garantir_lote_conferencia_e_itens(cursor, numero_lote):
    numero_lote = str(numero_lote or "").strip()
    if not numero_lote:
        return

    cursor.execute("SELECT * FROM lotes_envio WHERE numero_lote = ?", (numero_lote,))
    lote_envio = cursor.fetchone()

    tipo_lote = ""
    data_criacao = agora_str()
    if lote_envio:
        tipo_lote = str(lote_envio["tipo_lote"] or "")
        data_criacao = str(lote_envio["data_criacao"] or "") or data_criacao

    cursor.execute("SELECT numero_lote FROM lotes_conferencia WHERE numero_lote = ?", (numero_lote,))
    if not cursor.fetchone():
        cursor.execute("""
            INSERT INTO lotes_conferencia (numero_lote, tipo_lote, status, data_criacao)
            VALUES (?, ?, 'PENDENTE', ?)
        """, (numero_lote, tipo_lote, data_criacao))

    cursor.execute("SELECT COUNT(*) FROM lotes_itens WHERE numero_lote = ?", (numero_lote,))
    total_itens = int(cursor.fetchone()[0] or 0)
    if total_itens > 0:
        return

    cursor.execute("""
        SELECT numero_lote, codigo, sku, titulo, quantidade, endereco, lote_filete
        FROM lotes_envio_itens_snapshot
        WHERE numero_lote = ?
        ORDER BY endereco, sku, codigo
    """, (numero_lote,))
    itens_snapshot = cursor.fetchall()

    if not itens_snapshot:
        cursor.execute("""
            SELECT numero_lote, codigo, sku, endereco, titulo, quantidade, observacao
            FROM lotes_picking_itens
            WHERE numero_lote = ?
            ORDER BY endereco, sku, codigo
        """, (numero_lote,))
        itens_picking = cursor.fetchall()

        for item in itens_picking:
            cursor.execute("""
                INSERT OR IGNORE INTO lotes_itens (
                    numero_lote, codigo, sku, titulo, quantidade_esperada, endereco, lote_filete
                )
                VALUES (?, ?, ?, ?, ?, ?, ?)
            """, (
                item["numero_lote"],
                item["codigo"],
                item["sku"],
                item["titulo"],
                int(item["quantidade"] or 0),
                item["endereco"],
                ""
            ))
        return

    for item in itens_snapshot:
        cursor.execute("""
            INSERT OR IGNORE INTO lotes_itens (
                numero_lote, codigo, sku, titulo, quantidade_esperada, endereco, lote_filete
            )
            VALUES (?, ?, ?, ?, ?, ?, ?)
        """, (
            item["numero_lote"],
            item["codigo"],
            item["sku"],
            item["titulo"],
            int(item["quantidade"] or 0),
            item["endereco"],
            item["lote_filete"]
        ))


def sincronizar_embalagem_itens(cursor, numero_lote):
    garantir_tabela_embalagem(cursor)

    cursor.execute("""
        SELECT
            li.numero_lote,
            li.codigo,
            li.sku,
            li.titulo,
            li.quantidade_esperada,
            li.endereco,
            li.lote_filete,
            COALESCE(ci.observacao, '') AS observacao_conferencia
        FROM lotes_itens li
        INNER JOIN conferencia_itens ci
            ON li.numero_lote = ci.numero_lote AND li.codigo = ci.codigo
        WHERE li.numero_lote = ?
          AND COALESCE(ci.status_item, '') IN ('OK', 'DIVERGENTE')
        ORDER BY li.endereco, li.sku, li.codigo
    """, (numero_lote,))
    itens = cursor.fetchall()

    for item in itens:
        cursor.execute("""
            INSERT INTO lotes_embalagem_itens (
                numero_lote, codigo, sku, titulo, quantidade, endereco, lote_filete, observacao
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            ON CONFLICT(numero_lote, codigo) DO UPDATE SET
                sku = excluded.sku,
                titulo = excluded.titulo,
                quantidade = excluded.quantidade,
                endereco = excluded.endereco,
                lote_filete = excluded.lote_filete
        """, (
            item["numero_lote"],
            item["codigo"],
            item["sku"],
            item["titulo"],
            int(item["quantidade_esperada"] or 0),
            item["endereco"],
            item["lote_filete"],
            item["observacao_conferencia"]
        ))


def obter_nome_usuario_atual():
    usuario_atual = get_current_user()
    try:
        return str(getattr(usuario_atual, "nome", "") or getattr(usuario_atual, "email", "") or "").strip()
    except:
        return ""


def garantir_lote_conferencia_aberto(cursor, numero_lote):
    numero_lote = str(numero_lote or "").strip()
    if not numero_lote:
        return

    cursor.execute("SELECT * FROM lotes_envio WHERE numero_lote = ?", (numero_lote,))
    lote_envio = cursor.fetchone()

    tipo_lote = ""
    data_criacao = agora_str()
    if lote_envio:
        try:
            tipo_lote = str(lote_envio["tipo_lote"] or "")
            data_criacao = str(lote_envio["data_criacao"] or "") or data_criacao
        except:
            pass

    cursor.execute("SELECT numero_lote FROM lotes_conferencia WHERE numero_lote = ?", (numero_lote,))
    if not cursor.fetchone():
        cursor.execute("""
            INSERT INTO lotes_conferencia (numero_lote, tipo_lote, status, data_criacao)
            VALUES (?, ?, 'PENDENTE', ?)
        """, (numero_lote, tipo_lote, data_criacao))


def mover_item_picking_para_conferencia(cursor, numero_lote, codigo):
    numero_lote = str(numero_lote or "").strip()
    codigo = str(codigo or "").strip()
    if not numero_lote or not codigo:
        return

    garantir_lote_conferencia_aberto(cursor, numero_lote)

    cursor.execute("""
        INSERT OR IGNORE INTO lotes_itens (
            numero_lote, codigo, sku, titulo, quantidade_esperada, endereco, lote_filete
        )
        SELECT
            numero_lote,
            codigo,
            sku,
            titulo,
            quantidade,
            endereco,
            ''
        FROM lotes_picking_itens
        WHERE numero_lote = ? AND codigo = ?
    """, (numero_lote, codigo))


def mover_item_conferencia_para_embalagem(cursor, numero_lote, codigo):
    numero_lote = str(numero_lote or "").strip()
    codigo = str(codigo or "").strip()
    if not numero_lote or not codigo:
        return

    garantir_tabela_embalagem(cursor)

    cursor.execute("""
        INSERT INTO lotes_embalagem_itens (
            numero_lote, codigo, sku, titulo, quantidade, endereco, lote_filete, observacao
        )
        SELECT
            li.numero_lote,
            li.codigo,
            li.sku,
            li.titulo,
            li.quantidade_esperada,
            li.endereco,
            li.lote_filete,
            COALESCE(ci.observacao, '')
        FROM lotes_itens li
        LEFT JOIN conferencia_itens ci
            ON li.numero_lote = ci.numero_lote AND li.codigo = ci.codigo
        WHERE li.numero_lote = ? AND li.codigo = ?
        ON CONFLICT(numero_lote, codigo) DO UPDATE SET
            sku = excluded.sku,
            titulo = excluded.titulo,
            quantidade = excluded.quantidade,
            endereco = excluded.endereco,
            lote_filete = excluded.lote_filete,
            observacao = CASE
                WHEN COALESCE(lotes_embalagem_itens.observacao, '') = '' THEN excluded.observacao
                ELSE lotes_embalagem_itens.observacao
            END
    """, (numero_lote, codigo))


@app.route("/conferencia")
def conferencia():
    conn = sqlite3.connect("status.db")
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()

    cursor.execute("SELECT DISTINCT numero_lote FROM lotes_itens")
    for row_envio in cursor.fetchall():
        garantir_lote_conferencia_aberto(cursor, row_envio["numero_lote"])
    conn.commit()

    cursor.execute("""
        SELECT
            lc.numero_lote,
            lc.tipo_lote,
            lc.status,
            lc.data_criacao,
            COUNT(li.id) AS total_itens,
            SUM(CASE WHEN ci.status_item = 'OK' THEN 1 ELSE 0 END) AS itens_ok,
            SUM(CASE WHEN ci.status_item = 'DIVERGENTE' THEN 1 ELSE 0 END) AS itens_divergentes,
            SUM(CASE WHEN ci.id IS NOT NULL THEN 1 ELSE 0 END) AS itens_conferidos
        FROM lotes_conferencia lc
        LEFT JOIN lotes_itens li ON lc.numero_lote = li.numero_lote
        LEFT JOIN conferencia_itens ci
            ON lc.numero_lote = ci.numero_lote AND li.codigo = ci.codigo
        GROUP BY lc.numero_lote, lc.tipo_lote, lc.status, lc.data_criacao
        ORDER BY lc.data_criacao DESC
    """)

    lotes = cursor.fetchall()
    conn.close()

    return render_template("conferencia.html", lotes=lotes, lote=None, itens=[])


@app.route("/conferencia/<numero_lote>")
def conferencia_lote(numero_lote):
    conn = sqlite3.connect("status.db")
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    garantir_lote_conferencia_e_itens(cursor, numero_lote)
    conn.commit()

    cursor.execute("""
        SELECT
            lc.numero_lote,
            lc.tipo_lote,
            lc.status,
            lc.data_criacao,
            COUNT(li.id) AS total_itens,
            SUM(CASE WHEN ci.status_item = 'OK' THEN 1 ELSE 0 END) AS itens_ok,
            SUM(CASE WHEN ci.status_item = 'DIVERGENTE' THEN 1 ELSE 0 END) AS itens_divergentes,
            SUM(CASE WHEN ci.id IS NOT NULL THEN 1 ELSE 0 END) AS itens_conferidos
        FROM lotes_conferencia lc
        LEFT JOIN lotes_itens li ON lc.numero_lote = li.numero_lote
        LEFT JOIN conferencia_itens ci
            ON lc.numero_lote = ci.numero_lote AND li.codigo = ci.codigo
        WHERE lc.numero_lote = ?
        GROUP BY lc.numero_lote, lc.tipo_lote, lc.status, lc.data_criacao
    """, (numero_lote,))
    lote = cursor.fetchone()

    cursor.execute("""
        SELECT
            li.numero_lote,
            li.codigo,
            li.sku,
            li.titulo,
            li.quantidade_esperada,
            li.endereco,
            li.lote_filete,
            ci.quantidade_conferida,
            ci.foto_path,
            ci.status_item,
            ci.observacao,
            ci.conferido_em
        FROM lotes_itens li
        LEFT JOIN conferencia_itens ci
            ON li.numero_lote = ci.numero_lote AND li.codigo = ci.codigo
        WHERE li.numero_lote = ?
        ORDER BY li.endereco, li.sku
    """, (numero_lote,))
    itens = cursor.fetchall()

    cursor.execute("""
        SELECT
            lc.numero_lote,
            lc.tipo_lote,
            lc.status,
            lc.data_criacao,
            COUNT(li.id) AS total_itens,
            SUM(CASE WHEN ci.status_item = 'OK' THEN 1 ELSE 0 END) AS itens_ok,
            SUM(CASE WHEN ci.status_item = 'DIVERGENTE' THEN 1 ELSE 0 END) AS itens_divergentes,
            SUM(CASE WHEN ci.id IS NOT NULL THEN 1 ELSE 0 END) AS itens_conferidos
        FROM lotes_conferencia lc
        LEFT JOIN lotes_itens li ON lc.numero_lote = li.numero_lote
        LEFT JOIN conferencia_itens ci
            ON lc.numero_lote = ci.numero_lote AND li.codigo = ci.codigo
        GROUP BY lc.numero_lote, lc.tipo_lote, lc.status, lc.data_criacao
        ORDER BY lc.data_criacao DESC
    """)
    lotes = cursor.fetchall()

    conn.close()

    return render_template("conferencia.html", lotes=lotes, lote=lote, itens=itens)


@app.route("/salvar-conferencia-item", methods=["POST"])
def salvar_conferencia_item():
    numero_lote = request.form.get("numero_lote", "").strip()
    codigo = request.form.get("codigo", "").strip()
    sku = request.form.get("sku", "").strip()
    observacao = request.form.get("observacao", "").strip()

    try:
        quantidade_conferida = int(request.form.get("quantidade_conferida", 0))
    except:
        quantidade_conferida = 0

    foto = request.files.get("foto")
    foto_path = None

    conn = sqlite3.connect("status.db")
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()

    cursor.execute("""
        SELECT foto_path
        FROM conferencia_itens
        WHERE numero_lote = ? AND codigo = ?
    """, (numero_lote, codigo))
    registro_existente = cursor.fetchone()

    foto_path_existente = None
    if registro_existente:
        foto_path_existente = registro_existente["foto_path"]

    if foto and foto.filename:
        nome_original = secure_filename(foto.filename)
        timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
        nome_arquivo = f"{numero_lote}_{codigo}_{timestamp}_{nome_original}"
        caminho_arquivo = os.path.join(app.config["UPLOAD_FOLDER_CONFERENCIA"], nome_arquivo)
        foto.save(caminho_arquivo)
        foto_path = os.path.join("static", "uploads", "conferencia", nome_arquivo).replace("\\", "/")
    else:
        foto_path = foto_path_existente

    cursor.execute("""
        SELECT quantidade_esperada
        FROM lotes_itens
        WHERE numero_lote = ? AND codigo = ?
    """, (numero_lote, codigo))
    item_lote = cursor.fetchone()

    quantidade_esperada = 0
    if item_lote:
        try:
            quantidade_esperada = int(item_lote["quantidade_esperada"] or 0)
        except:
            quantidade_esperada = 0

    if quantidade_conferida <= 0:
        status_item = "PENDENTE"
    elif quantidade_conferida != quantidade_esperada:
        status_item = "DIVERGENTE"
    else:
        status_item = "OK"

    conferido_em = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    cursor.execute("""
        INSERT INTO conferencia_itens (
            numero_lote,
            codigo,
            sku,
            quantidade_conferida,
            foto_path,
            status_item,
            observacao,
            conferido_em
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        ON CONFLICT(numero_lote, codigo) DO UPDATE SET
            sku=excluded.sku,
            quantidade_conferida=excluded.quantidade_conferida,
            foto_path=excluded.foto_path,
            status_item=excluded.status_item,
            observacao=excluded.observacao,
            conferido_em=excluded.conferido_em
    """, (
        numero_lote,
        codigo,
        sku,
        quantidade_conferida,
        foto_path,
        status_item,
        observacao,
        conferido_em
    ))

    if status_item in ["OK", "DIVERGENTE"]:
        mover_item_conferencia_para_embalagem(cursor, numero_lote, codigo)

    cursor.execute("""
        SELECT COUNT(*)
        FROM lotes_itens
        WHERE numero_lote = ?
    """, (numero_lote,))
    total_itens = cursor.fetchone()[0]

    cursor.execute("""
        SELECT COUNT(*)
        FROM conferencia_itens
        WHERE numero_lote = ?
    """, (numero_lote,))
    total_conferidos = cursor.fetchone()[0]

    novo_status_lote = "PENDENTE"
    data_fechamento = None

    if total_itens > 0 and total_conferidos >= total_itens:
        cursor.execute("""
            SELECT COUNT(*)
            FROM conferencia_itens
            WHERE numero_lote = ? AND status_item = 'DIVERGENTE'
        """, (numero_lote,))
        qtd_divergentes = cursor.fetchone()[0]

        if qtd_divergentes > 0:
            novo_status_lote = "CONFERIDO COM DIVERGÊNCIA"
        else:
            novo_status_lote = "CONFERIDO"

        data_fechamento = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    cursor.execute("""
        UPDATE lotes_conferencia
        SET status = ?, data_fechamento = ?
        WHERE numero_lote = ?
    """, (novo_status_lote, data_fechamento, numero_lote))

    if total_itens > 0 and total_conferidos >= total_itens:
        sincronizar_embalagem_itens(cursor, numero_lote)

    conn.commit()
    conn.close()

    return redirect(f"/conferencia/{numero_lote}")


def init_db():
    conn = sqlite3.connect("status.db")
    cursor = conn.cursor()

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS status_cards (
            codigo TEXT PRIMARY KEY,
            status TEXT,
            quantidade INTEGER DEFAULT 0,
            estrategia TEXT DEFAULT '',
            prioridade TEXT DEFAULT ''
        )
    """)

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS comentarios (
            sku TEXT PRIMARY KEY,
            comentario TEXT
        )
    """)

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS comentarios_mlb (
            codigo TEXT PRIMARY KEY,
            comentario TEXT
        )
    """)

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS comentarios_mlb_chat (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            codigo TEXT NOT NULL,
            mensagem TEXT NOT NULL,
            data_hora TEXT
        )
    """)

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS lotes_conferencia (
            numero_lote TEXT PRIMARY KEY,
            tipo_lote TEXT,
            status TEXT DEFAULT 'PENDENTE',
            data_criacao TEXT,
            data_fechamento TEXT
        )
    """)

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS lotes_itens (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            numero_lote TEXT,
            codigo TEXT,
            sku TEXT,
            titulo TEXT,
            quantidade_esperada INTEGER DEFAULT 0,
            endereco TEXT,
            lote_filete TEXT,
            UNIQUE(numero_lote, codigo)
        )
    """)

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS conferencia_itens (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            numero_lote TEXT,
            codigo TEXT,
            sku TEXT,
            quantidade_conferida INTEGER DEFAULT 0,
            foto_path TEXT,
            status_item TEXT DEFAULT 'PENDENTE',
            observacao TEXT,
            conferido_em TEXT,
            UNIQUE(numero_lote, codigo)
        )
    """)

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS historico_filetes (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            numero_lote TEXT,
            tipo_lote TEXT,
            codigo TEXT,
            sku TEXT,
            titulo TEXT,
            nickname TEXT,
            quantidade INTEGER DEFAULT 0,
            endereco TEXT,
            lote_filete TEXT,
            data_geracao TEXT
        )
    """)

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS lotes_envio (
            numero_lote TEXT PRIMARY KEY,
            tipo_lote TEXT,
            total_mlbs INTEGER DEFAULT 0,
            total_pecas INTEGER DEFAULT 0,
            status TEXT DEFAULT 'CRIADO',
            responsavel TEXT DEFAULT '',
            transportadora TEXT DEFAULT '',
            observacao TEXT DEFAULT '',
            prioridade TEXT DEFAULT '',
            data_envio TEXT DEFAULT '',
            status_expedicao TEXT DEFAULT 'AGUARDANDO',
            status_ecommerce TEXT DEFAULT 'AGUARDANDO',
            origem TEXT DEFAULT 'MANUAL',
            data_criacao TEXT
        )
    """)

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS lotes_envio_itens_snapshot (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            numero_lote TEXT,
            tipo_lote TEXT,
            codigo TEXT,
            sku TEXT,
            titulo TEXT,
            nickname TEXT,
            quantidade INTEGER DEFAULT 0,
            endereco TEXT,
            lote_filete TEXT,
            estrategia TEXT DEFAULT '',
            motivo_envio TEXT DEFAULT '',
            comentario_mlb TEXT DEFAULT '',
            dados_json TEXT,
            data_geracao TEXT
        )
    """)

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS lotes_picking_itens (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            numero_lote TEXT,
            codigo TEXT,
            sku TEXT,
            endereco TEXT,
            titulo TEXT DEFAULT '',
            conta TEXT DEFAULT '',
            selo TEXT DEFAULT '',
            quantidade INTEGER DEFAULT 0,
            observacao TEXT DEFAULT '',
            coletado INTEGER DEFAULT 0,
            coletado_em TEXT DEFAULT '',
            quantidade_informada INTEGER DEFAULT 0,
            divergencia INTEGER DEFAULT 0,
            divergencia_em TEXT DEFAULT '',
            UNIQUE(numero_lote, codigo)
        )
    """)

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS lotes_embalagem_itens (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            numero_lote TEXT,
            codigo TEXT,
            sku TEXT,
            titulo TEXT DEFAULT '',
            quantidade INTEGER DEFAULT 0,
            endereco TEXT DEFAULT '',
            lote_filete TEXT DEFAULT '',
            observacao TEXT DEFAULT '',
            embalado INTEGER DEFAULT 0,
            embalado_em TEXT DEFAULT '',
            embalado_por TEXT DEFAULT '',
            quantidade_embalada INTEGER DEFAULT 0,
            divergencia INTEGER DEFAULT 0,
            divergencia_em TEXT DEFAULT '',
            UNIQUE(numero_lote, codigo)
        )
    """)

    cursor.execute("PRAGMA table_info(status_cards)")
    colunas_status = [col[1] for col in cursor.fetchall()]

    if "motivo_envio" not in colunas_status:
        cursor.execute("ALTER TABLE status_cards ADD COLUMN motivo_envio TEXT DEFAULT ''")

    if "quantidade" not in colunas_status:
        cursor.execute("ALTER TABLE status_cards ADD COLUMN quantidade INTEGER DEFAULT 0")

    if "estrategia" not in colunas_status:
        cursor.execute("ALTER TABLE status_cards ADD COLUMN estrategia TEXT DEFAULT ''")

    if "prioridade" not in colunas_status:
        cursor.execute("ALTER TABLE status_cards ADD COLUMN prioridade TEXT DEFAULT ''")

    cursor.execute("PRAGMA table_info(lotes_envio)")
    colunas_lotes_envio = [col[1] for col in cursor.fetchall()]

    cursor.execute("PRAGMA table_info(lotes_picking_itens)")
    colunas_picking = [col[1] for col in cursor.fetchall()]

    if "titulo" not in colunas_picking:
        cursor.execute("ALTER TABLE lotes_picking_itens ADD COLUMN titulo TEXT DEFAULT ''")

    if "conta" not in colunas_picking:
        cursor.execute("ALTER TABLE lotes_picking_itens ADD COLUMN conta TEXT DEFAULT ''")

    if "selo" not in colunas_picking:
        cursor.execute("ALTER TABLE lotes_picking_itens ADD COLUMN selo TEXT DEFAULT ''")

    if "observacao" not in colunas_picking:
        cursor.execute("ALTER TABLE lotes_picking_itens ADD COLUMN observacao TEXT DEFAULT ''")

    if "quantidade_informada" not in colunas_picking:
        cursor.execute("ALTER TABLE lotes_picking_itens ADD COLUMN quantidade_informada INTEGER DEFAULT 0")

    if "divergencia" not in colunas_picking:
        cursor.execute("ALTER TABLE lotes_picking_itens ADD COLUMN divergencia INTEGER DEFAULT 0")

    if "divergencia_em" not in colunas_picking:
        cursor.execute("ALTER TABLE lotes_picking_itens ADD COLUMN divergencia_em TEXT DEFAULT ''")

    if "tipo_lote" not in colunas_lotes_envio:
        cursor.execute("ALTER TABLE lotes_envio ADD COLUMN tipo_lote TEXT")

    if "total_mlbs" not in colunas_lotes_envio:
        cursor.execute("ALTER TABLE lotes_envio ADD COLUMN total_mlbs INTEGER DEFAULT 0")

    if "total_pecas" not in colunas_lotes_envio:
        cursor.execute("ALTER TABLE lotes_envio ADD COLUMN total_pecas INTEGER DEFAULT 0")

    if "status" not in colunas_lotes_envio:
        cursor.execute("ALTER TABLE lotes_envio ADD COLUMN status TEXT DEFAULT 'CRIADO'")

    if "responsavel" not in colunas_lotes_envio:
        cursor.execute("ALTER TABLE lotes_envio ADD COLUMN responsavel TEXT DEFAULT ''")

    if "transportadora" not in colunas_lotes_envio:
        cursor.execute("ALTER TABLE lotes_envio ADD COLUMN transportadora TEXT DEFAULT ''")

    if "observacao" not in colunas_lotes_envio:
        cursor.execute("ALTER TABLE lotes_envio ADD COLUMN observacao TEXT DEFAULT ''")

    if "prioridade" not in colunas_lotes_envio:
        cursor.execute("ALTER TABLE lotes_envio ADD COLUMN prioridade TEXT DEFAULT ''")

    if "data_envio" not in colunas_lotes_envio:
        cursor.execute("ALTER TABLE lotes_envio ADD COLUMN data_envio TEXT DEFAULT ''")

    if "status_expedicao" not in colunas_lotes_envio:
        cursor.execute("ALTER TABLE lotes_envio ADD COLUMN status_expedicao TEXT DEFAULT 'AGUARDANDO'")

    if "status_ecommerce" not in colunas_lotes_envio:
        cursor.execute("ALTER TABLE lotes_envio ADD COLUMN status_ecommerce TEXT DEFAULT 'AGUARDANDO'")

    if "origem" not in colunas_lotes_envio:
        cursor.execute("ALTER TABLE lotes_envio ADD COLUMN origem TEXT DEFAULT 'MANUAL'")

    if "data_criacao" not in colunas_lotes_envio:
        cursor.execute("ALTER TABLE lotes_envio ADD COLUMN data_criacao TEXT")

    if "timeline_json" not in colunas_lotes_envio:
        cursor.execute("ALTER TABLE lotes_envio ADD COLUMN timeline_json TEXT DEFAULT ''")

    if "etapa_atual" not in colunas_lotes_envio:
        cursor.execute("ALTER TABLE lotes_envio ADD COLUMN etapa_atual TEXT DEFAULT 'MLBS EM ANÁLISE'")

    if "data_coleta_agendada" not in colunas_lotes_envio:
        cursor.execute("ALTER TABLE lotes_envio ADD COLUMN data_coleta_agendada TEXT DEFAULT ''")

    conn.commit()
    conn.close()


@app.route("/")
def index():
    return render_template("index.html")


@app.route("/home")
def home():
    return render_template("home.html")

def garantir_tabela_lotes_envio_snapshot():
    conn = sqlite3.connect("status.db")
    cursor = conn.cursor()

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS lotes_envio_itens_snapshot (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            numero_lote TEXT,
            tipo_lote TEXT,
            codigo TEXT,
            sku TEXT,
            titulo TEXT,
            nickname TEXT,
            quantidade INTEGER DEFAULT 0,
            endereco TEXT,
            lote_filete TEXT,
            estrategia TEXT DEFAULT '',
            motivo_envio TEXT DEFAULT '',
            comentario_mlb TEXT DEFAULT '',
            dados_json TEXT,
            data_geracao TEXT
        )
    """)

    conn.commit()
    conn.close()


def montar_mapa_lotes_exportacao(cursor):
    mapa = {}

    consultas = [
        """
        SELECT
            s.codigo,
            s.numero_lote,
            s.lote_filete,
            COALESCE(le.data_coleta_agendada, '') AS data_coleta_agendada
        FROM lotes_envio_itens_snapshot s
        INNER JOIN (
            SELECT codigo, MAX(id) AS max_id
            FROM lotes_envio_itens_snapshot
            GROUP BY codigo
        ) ult
            ON ult.codigo = s.codigo
           AND ult.max_id = s.id
        LEFT JOIN lotes_envio le
            ON le.numero_lote = s.numero_lote
        """,
        """
        SELECT
            hf.codigo,
            hf.numero_lote,
            hf.lote_filete,
            COALESCE(le.data_coleta_agendada, '') AS data_coleta_agendada
        FROM historico_filetes hf
        INNER JOIN (
            SELECT codigo, MAX(id) AS max_id
            FROM historico_filetes
            GROUP BY codigo
        ) ult
            ON ult.codigo = hf.codigo
           AND ult.max_id = hf.id
        LEFT JOIN lotes_envio le
            ON le.numero_lote = hf.numero_lote
        """
    ]

    for consulta in consultas:
        try:
            cursor.execute(consulta)
            for codigo, numero_lote, lote_filete, data_coleta_agendada in cursor.fetchall():
                codigo = str(codigo or "").strip()
                if not codigo or codigo in mapa:
                    continue
                mapa[codigo] = {
                    "numero_lote": str(numero_lote or "").strip(),
                    "lote": str(lote_filete or "").strip(),
                    "data_coleta": str(data_coleta_agendada or "").strip()
                }
        except:
            pass

        if mapa:
            break

    return mapa


def preparar_dataframe_exportacao(df, mapa_lotes=None, numero_lote_fixo="", data_coleta_fixa=""):
    df = df.copy().fillna("")

    if "Código do Anúncio" not in df.columns:
        df["Código do Anúncio"] = ""

    codigo_series = df["Código do Anúncio"].astype(str).str.strip()

    if mapa_lotes is None:
        mapa_lotes = {}

    mapa_numero_lote = {str(codigo): str((dados or {}).get("numero_lote", "") or "").strip() for codigo, dados in mapa_lotes.items()}
    mapa_nome_lote = {str(codigo): str((dados or {}).get("lote", "") or "").strip() for codigo, dados in mapa_lotes.items()}
    mapa_data_coleta = {str(codigo): str((dados or {}).get("data_coleta", "") or "").strip() for codigo, dados in mapa_lotes.items()}

    if "Nickname" not in df.columns:
        df["Nickname"] = ""

    if "SKU" not in df.columns:
        df["SKU"] = ""

    if "Quantidade para Enviar" not in df.columns:
        df["Quantidade para Enviar"] = 0
    df["Quantidade para Enviar"] = pd.to_numeric(df["Quantidade para Enviar"], errors="coerce").fillna(0).astype(int)

    if "ESTOQUE TOTAL SIGNUS" not in df.columns:
        df["ESTOQUE TOTAL SIGNUS"] = ""

    estoque_numerico = df["ESTOQUE TOTAL SIGNUS"].apply(numero_float)
    quantidade_numerica = pd.to_numeric(df["Quantidade para Enviar"], errors="coerce").fillna(0)
    df["APÓS ENVIO"] = (estoque_numerico - quantidade_numerica).astype(int)

    if "RETORNO" not in df.columns:
        df["RETORNO"] = ""

    if "ENDEREÇO" not in df.columns:
        df["ENDEREÇO"] = ""

    if "LOTE" not in df.columns:
        df["LOTE"] = ""
    df["LOTE"] = df["LOTE"].where(df["LOTE"].astype(str).str.strip() != "", codigo_series.map(mapa_nome_lote).fillna(""))

    if "NUMERO DO LOTE" not in df.columns:
        df["NUMERO DO LOTE"] = ""
    if numero_lote_fixo:
        df["NUMERO DO LOTE"] = str(numero_lote_fixo)
    else:
        df["NUMERO DO LOTE"] = df["NUMERO DO LOTE"].where(
            df["NUMERO DO LOTE"].astype(str).str.strip() != "",
            codigo_series.map(mapa_numero_lote).fillna("")
        )

    if "DATA DA COLETA" not in df.columns:
        df["DATA DA COLETA"] = ""
    if data_coleta_fixa:
        df["DATA DA COLETA"] = str(data_coleta_fixa)
    else:
        df["DATA DA COLETA"] = df["DATA DA COLETA"].where(
            df["DATA DA COLETA"].astype(str).str.strip() != "",
            codigo_series.map(mapa_data_coleta).fillna("")
        )

    if "PRODUTO" not in df.columns:
        df["PRODUTO"] = ""
    if "Título" in df.columns:
        df["PRODUTO"] = df["PRODUTO"].where(df["PRODUTO"].astype(str).str.strip() != "", df["Título"].astype(str))

    if "Motivo" not in df.columns:
        df["Motivo"] = ""
    if "Motivo do Envio" in df.columns:
        df["Motivo"] = df["Motivo"].where(df["Motivo"].astype(str).str.strip() != "", df["Motivo do Envio"].astype(str))

    colunas_exportar = [
        "Nickname",
        "Código do Anúncio",
        "SKU",
        "Quantidade para Enviar",
        "ESTOQUE TOTAL SIGNUS",
        "APÓS ENVIO",
        "ENDEREÇO",
        "LOTE",
        "PRODUTO",
        "Motivo",
        "RETORNO",
        "NUMERO DO LOTE",
        "DATA DA COLETA"
    ]

    for col in colunas_exportar:
        if col not in df.columns:
            df[col] = ""

    return df[colunas_exportar].copy()


def obter_ultimo_lote_por_codigo(cursor):
    mapa = {}

    try:
        cursor.execute("""
            SELECT
                s.codigo,
                s.numero_lote,
                s.quantidade,
                s.data_geracao,
                COALESCE(le.status, '') AS status_lote
            FROM lotes_envio_itens_snapshot s
            INNER JOIN (
                SELECT codigo, MAX(id) AS max_id
                FROM lotes_envio_itens_snapshot
                GROUP BY codigo
            ) ult
                ON ult.codigo = s.codigo
               AND ult.max_id = s.id
            LEFT JOIN lotes_envio le
                ON le.numero_lote = s.numero_lote
        """)

        for codigo, numero_lote, quantidade, data_geracao, status_lote in cursor.fetchall():
            mapa[str(codigo)] = {
                "numero_lote": str(numero_lote or ""),
                "quantidade": int(quantidade or 0),
                "data_geracao": str(data_geracao or ""),
                "status_lote": str(status_lote or "")
            }
    except:
        pass

    if mapa:
        return mapa

    try:
        cursor.execute("""
            SELECT
                hf.codigo,
                hf.numero_lote,
                hf.quantidade,
                hf.data_geracao,
                COALESCE(le.status, '') AS status_lote
            FROM historico_filetes hf
            INNER JOIN (
                SELECT codigo, MAX(id) AS max_id
                FROM historico_filetes
                GROUP BY codigo
            ) ult
                ON ult.codigo = hf.codigo
               AND ult.max_id = hf.id
            LEFT JOIN lotes_envio le
                ON le.numero_lote = hf.numero_lote
        """)

        for codigo, numero_lote, quantidade, data_geracao, status_lote in cursor.fetchall():
            mapa[str(codigo)] = {
                "numero_lote": str(numero_lote or ""),
                "quantidade": int(quantidade or 0),
                "data_geracao": str(data_geracao or ""),
                "status_lote": str(status_lote or "")
            }
    except:
        pass

    return mapa

@app.route("/metricas-full")
def metricas_full():
    garantir_tabela_lotes_envio_snapshot()

    conn = sqlite3.connect("status.db")
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()

    cursor.execute("""
        SELECT *
        FROM lotes_envio
        ORDER BY data_criacao DESC, numero_lote DESC
    """)
    lotes_rows = cursor.fetchall()

    cursor.execute("""
        SELECT *
        FROM lotes_envio_itens_snapshot
        ORDER BY numero_lote DESC, endereco, sku, codigo
    """)
    itens_snapshot = cursor.fetchall()

    if not itens_snapshot:
        cursor.execute("""
            SELECT
                li.numero_lote,
                COALESCE(le.tipo_lote, lc.tipo_lote, '') AS tipo_lote,
                li.codigo,
                li.sku,
                li.titulo,
                '' AS nickname,
                li.quantidade_esperada AS quantidade,
                li.endereco,
                li.lote_filete,
                '' AS estrategia,
                '' AS motivo_envio,
                '' AS comentario_mlb,
                '' AS dados_json,
                COALESCE(le.data_criacao, lc.data_criacao, '') AS data_geracao
            FROM lotes_itens li
            LEFT JOIN lotes_envio le
                ON le.numero_lote = li.numero_lote
            LEFT JOIN lotes_conferencia lc
                ON lc.numero_lote = li.numero_lote
            ORDER BY li.numero_lote DESC, li.endereco, li.sku, li.codigo
        """)
        itens_snapshot = cursor.fetchall()

    itens_snapshot = enriquecer_itens_lote_com_base([dict(item) for item in itens_snapshot])

    lotes_itens_map = {}
    for item in itens_snapshot:
        numero_lote = item["numero_lote"]
        lotes_itens_map.setdefault(numero_lote, []).append(item)

    total_lotes = len(lotes_rows)
    total_mlbs = sum(int(lote["total_mlbs"] or 0) for lote in lotes_rows)
    total_pecas = sum(int(lote["total_pecas"] or 0) for lote in lotes_rows)

    contas_total = {}
    for item in itens_snapshot:
        conta = str(item["nickname"] or "SEM CONTA").strip() or "SEM CONTA"
        if conta not in contas_total:
            contas_total[conta] = {"mlbs": set(), "pecas": 0}
        contas_total[conta]["mlbs"].add(str(item["codigo"] or ""))
        try:
            contas_total[conta]["pecas"] += int(item["quantidade"] or 0)
        except:
            pass

    contas_resumo = []
    for conta, dados in sorted(contas_total.items(), key=lambda x: x[0]):
        contas_resumo.append({
            "conta": conta,
            "mlbs": len(dados["mlbs"]),
            "pecas": int(dados["pecas"])
        })

    lotes = []
    total_abertos = 0
    total_fechados = 0
    soma_lead_time = 0
    lotes_com_lead_time = 0

    for lote in lotes_rows:
        lote_dict = dict(lote)
        numero_lote = lote_dict["numero_lote"]
        itens_lote = lotes_itens_map.get(numero_lote, [])
        timeline = carregar_timeline_json(lote_dict.get("timeline_json", ""))
        etapa_atual = lote_dict.get("etapa_atual") or TIMELINE_ETAPAS[0]

        if not any(str(v or "").strip() for v in timeline.values()):
            criacao = lote_dict.get("data_criacao") or agora_str()
            timeline[etapa_atual] = criacao
            lote_dict["timeline_json"] = timeline_para_json(timeline)

        status_abertura = obter_status_abertura_por_etapa(etapa_atual)
        if status_abertura == "ABERTO":
            total_abertos += 1
        else:
            total_fechados += 1

        lead_time_segundos = calcular_lead_time_segundos(timeline)
        if lead_time_segundos > 0:
            soma_lead_time += lead_time_segundos
            lotes_com_lead_time += 1

        lote_dict["status"] = status_abertura
        lote_dict["etapa_atual"] = etapa_atual
        lote_dict["timeline_exibicao"] = construir_timeline_exibicao(timeline, etapa_atual)
        lote_dict["lead_time_segundos"] = lead_time_segundos
        lote_dict["lead_time_humano"] = formatar_duracao_humana(lead_time_segundos)
        lote_dict["data_criacao_br"] = formatar_data_hora_br(lote_dict.get("data_criacao"))
        lote_dict["data_coleta_agendada_br"] = formatar_data_br(lote_dict.get("data_coleta_agendada"))
        lote_dict["contas_resumo"] = resumir_contas_lote(itens_lote)
        lotes.append(lote_dict)

    lead_time_medio = formatar_duracao_humana(int(soma_lead_time / lotes_com_lead_time)) if lotes_com_lead_time else "-"

    conn.close()

    return render_template(
        "metricas_full.html",
        lotes=lotes,
        lotes_itens_map=lotes_itens_map,
        total_lotes=total_lotes,
        total_mlbs=total_mlbs,
        total_pecas=total_pecas,
        contas_resumo=contas_resumo,
        total_abertos=total_abertos,
        total_fechados=total_fechados,
        lead_time_medio=lead_time_medio,
        timeline_etapas=TIMELINE_ETAPAS,
        formatar_data_hora_br=formatar_data_hora_br
    )


def segundos_para_humano_operacional(segundos):
    try:
        segundos = int(segundos or 0)
    except:
        segundos = 0

    if segundos <= 0:
        return "-"

    horas = segundos // 3600
    minutos = (segundos % 3600) // 60
    seg = segundos % 60

    partes = []
    if horas:
        partes.append(f"{horas}h")
    if minutos:
        partes.append(f"{minutos}min")
    if seg and not horas:
        partes.append(f"{seg}s")

    return " ".join(partes) if partes else "0s"


def garantir_colunas_gestao_operacional(cursor):
    try:
        cursor.execute("PRAGMA table_info(lotes_picking_itens)")
        colunas = [col[1] for col in cursor.fetchall()]
        if "coletado_por" not in colunas:
            cursor.execute("ALTER TABLE lotes_picking_itens ADD COLUMN coletado_por TEXT DEFAULT ''")
        if "divergencia_por" not in colunas:
            cursor.execute("ALTER TABLE lotes_picking_itens ADD COLUMN divergencia_por TEXT DEFAULT ''")
    except:
        pass


@app.route("/gestao-operacional")
def gestao_operacional():
    conn = sqlite3.connect("status.db")
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()

    garantir_colunas_gestao_operacional(cursor)
    conn.commit()

    cursor.execute("""
        SELECT *
        FROM lotes_envio
        ORDER BY COALESCE(data_criacao, '') DESC, numero_lote DESC
    """)
    lotes_rows = cursor.fetchall()

    lotes = []
    lotes_andamento = []
    lotes_finalizados = []
    itens_geral = []

    total_lotes_andamento = 0
    total_lotes_finalizados = 0
    total_itens = 0
    total_coletados = 0
    total_divergencias = 0
    total_quantidade_coletada = 0
    soma_tempo_item = 0
    qtd_tempo_item = 0

    for lote_row in lotes_rows:
        lote = dict(lote_row)
        numero_lote = str(lote.get("numero_lote") or "").strip()
        etapa_atual = lote.get("etapa_atual") or TIMELINE_ETAPAS[0]
        status_abertura = obter_status_abertura_por_etapa(etapa_atual)

        cursor.execute("""
            SELECT
                numero_lote,
                codigo,
                sku,
                endereco,
                titulo,
                conta,
                selo,
                quantidade,
                observacao,
                coletado,
                coletado_em,
                quantidade_informada,
                divergencia,
                divergencia_em,
                COALESCE(coletado_por, '') AS coletado_por,
                COALESCE(divergencia_por, '') AS divergencia_por
            FROM lotes_picking_itens
            WHERE numero_lote = ?
            ORDER BY
                CASE
                    WHEN COALESCE(coletado_em, '') <> '' THEN coletado_em
                    WHEN COALESCE(divergencia_em, '') <> '' THEN divergencia_em
                    ELSE '9999-12-31 23:59:59'
                END,
                endereco,
                sku,
                codigo
        """, (numero_lote,))
        itens = [dict(row) for row in cursor.fetchall()]

        if not itens:
            itens_snapshot = carregar_itens_snapshot_lote(numero_lote)
            if itens_snapshot:
                try:
                    df_boot = pd.DataFrame(itens_snapshot)
                    sincronizar_picking_itens(numero_lote, df_boot)
                    garantir_colunas_gestao_operacional(cursor)
                    conn.commit()
                    cursor.execute("""
                        SELECT
                            numero_lote,
                            codigo,
                            sku,
                            endereco,
                            titulo,
                            conta,
                            selo,
                            quantidade,
                            observacao,
                            coletado,
                            coletado_em,
                            quantidade_informada,
                            divergencia,
                            divergencia_em,
                            COALESCE(coletado_por, '') AS coletado_por,
                            COALESCE(divergencia_por, '') AS divergencia_por
                        FROM lotes_picking_itens
                        WHERE numero_lote = ?
                        ORDER BY endereco, sku, codigo
                    """, (numero_lote,))
                    itens = [dict(row) for row in cursor.fetchall()]
                except:
                    itens = []

        qtd_total = len(itens)
        qtd_coletados = sum(1 for item in itens if int(item.get("coletado") or 0) == 1)
        qtd_divergencias = sum(1 for item in itens if int(item.get("divergencia") or 0) == 1)

        if qtd_total == 0:
            continue

        lote_finalizado = (qtd_coletados + qtd_divergencias) >= qtd_total

        total_itens += qtd_total
        total_coletados += qtd_coletados
        total_divergencias += qtd_divergencias

        eventos = []
        for item in itens:
            fim_raw = str(item.get("coletado_em") or item.get("divergencia_em") or "").strip()
            fim_dt = parse_data_hora(fim_raw)
            if fim_dt:
                eventos.append((fim_dt, item))

        eventos.sort(key=lambda x: x[0])

        primeiro_dt = eventos[0][0] if eventos else None
        ultimo_dt = eventos[-1][0] if eventos else None
        lote_segundos = int((ultimo_dt - primeiro_dt).total_seconds()) if primeiro_dt and ultimo_dt else 0

        item_rows = []
        for idx, (fim_dt, item) in enumerate(eventos):
            inicio_dt = eventos[idx - 1][0] if idx > 0 else None
            proximo_dt = eventos[idx + 1][0] if idx + 1 < len(eventos) else None

            tempo_item_seg = int((fim_dt - inicio_dt).total_seconds()) if inicio_dt else 0
            tempo_ate_proxima_seg = int((proximo_dt - fim_dt).total_seconds()) if proximo_dt else 0

            if tempo_item_seg > 0:
                soma_tempo_item += tempo_item_seg
                qtd_tempo_item += 1

            quantidade_coletada = int(item.get("quantidade_informada") or 0)
            if quantidade_coletada <= 0 and int(item.get("coletado") or 0) == 1:
                quantidade_coletada = int(item.get("quantidade") or 0)
            total_quantidade_coletada += quantidade_coletada

            coletor = str(item.get("coletado_por") or item.get("divergencia_por") or "").strip() or "Não informado"

            item_view = {
                "numero_lote": numero_lote,
                "codigo": item.get("codigo") or "",
                "sku": item.get("sku") or "-",
                "titulo": item.get("titulo") or "-",
                "endereco": item.get("endereco") or "-",
                "selo": item.get("selo") or "-",
                "conta": item.get("conta") or "-",
                "inicio": inicio_dt.strftime("%d/%m/%Y %H:%M:%S") if inicio_dt else "-",
                "fim": fim_dt.strftime("%d/%m/%Y %H:%M:%S"),
                "data_coleta": fim_dt.strftime("%d/%m/%Y"),
                "tempo_item": segundos_para_humano_operacional(tempo_item_seg),
                "tempo_ate_proxima": segundos_para_humano_operacional(tempo_ate_proxima_seg),
                "coletor": coletor,
                "quantidade_coletada": quantidade_coletada,
                "quantidade_esperada": int(item.get("quantidade") or 0),
                "status": "Divergência" if int(item.get("divergencia") or 0) == 1 else "Coletado",
                "observacao": item.get("observacao") or ""
            }
            item_rows.append(item_view)
            itens_geral.append(item_view)

        percentual = int(round(((qtd_coletados + qtd_divergencias) / qtd_total) * 100)) if qtd_total else 0

        lote_view = {
            "numero_lote": numero_lote,
            "tipo_lote": lote.get("tipo_lote") or "Diversos",
            "etapa_atual": etapa_atual,
            "status": "FINALIZADO" if lote_finalizado else status_abertura,
            "data_criacao": formatar_data_hora_br(lote.get("data_criacao")),
            "data_coleta_agendada": formatar_data_br(lote.get("data_coleta_agendada")),
            "total_itens": qtd_total,
            "coletados": qtd_coletados,
            "divergencias": qtd_divergencias,
            "pendentes": max(qtd_total - qtd_coletados - qtd_divergencias, 0),
            "percentual": percentual,
            "inicio_coleta": primeiro_dt.strftime("%d/%m/%Y %H:%M:%S") if primeiro_dt else "-",
            "fim_coleta": ultimo_dt.strftime("%d/%m/%Y %H:%M:%S") if ultimo_dt else "-",
            "tempo_total_coleta": segundos_para_humano_operacional(lote_segundos),
            "tempo_medio_item": segundos_para_humano_operacional(int(lote_segundos / max(len(eventos) - 1, 1))) if lote_segundos else "-",
            "itens": item_rows
        }

        lotes.append(lote_view)
        if lote_finalizado:
            lotes_finalizados.append(lote_view)
            total_lotes_finalizados += 1
        else:
            lotes_andamento.append(lote_view)
            total_lotes_andamento += 1

    tempo_medio_geral = segundos_para_humano_operacional(int(soma_tempo_item / qtd_tempo_item)) if qtd_tempo_item else "-"

    conn.close()

    return render_template(
        "gestao_operacional.html",
        lotes=lotes,
        lotes_andamento=lotes_andamento,
        lotes_finalizados=lotes_finalizados,
        itens_geral=itens_geral,
        total_lotes_andamento=total_lotes_andamento,
        total_lotes_finalizados=total_lotes_finalizados,
        total_itens=total_itens,
        total_coletados=total_coletados,
        total_divergencias=total_divergencias,
        total_pendentes=max(total_itens - total_coletados - total_divergencias, 0),
        total_quantidade_coletada=total_quantidade_coletada,
        tempo_medio_geral=tempo_medio_geral
    )


@app.route("/picking")
def picking_lista():
    conn = sqlite3.connect("status.db")
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()

    cursor.execute("""
        SELECT *
        FROM lotes_envio
        ORDER BY data_criacao DESC, numero_lote DESC
    """)
    lotes_rows = cursor.fetchall()

    lotes = []
    for lote in lotes_rows:
        lote_dict = dict(lote)
        etapa_atual = lote_dict.get("etapa_atual") or TIMELINE_ETAPAS[0]
        if obter_status_abertura_por_etapa(etapa_atual) != "ABERTO":
            continue

        numero_lote = lote_dict["numero_lote"]
        cursor.execute("SELECT COUNT(*) FROM lotes_picking_itens WHERE numero_lote = ?", (numero_lote,))
        total_itens = int(cursor.fetchone()[0] or 0)

        if total_itens == 0:
            itens_snapshot = carregar_itens_snapshot_lote(numero_lote)
            if itens_snapshot:
                df_boot = pd.DataFrame(itens_snapshot)
                sincronizar_picking_itens(numero_lote, df_boot)
                cursor.execute("SELECT COUNT(*) FROM lotes_picking_itens WHERE numero_lote = ?", (numero_lote,))
                total_itens = int(cursor.fetchone()[0] or 0)

        cursor.execute("SELECT COUNT(*) FROM lotes_picking_itens WHERE numero_lote = ? AND coletado = 1", (numero_lote,))
        itens_coletados = int(cursor.fetchone()[0] or 0)

        lote_dict["etapa_atual"] = etapa_atual
        lote_dict["status"] = obter_status_abertura_por_etapa(etapa_atual)
        lote_dict["total_itens"] = total_itens
        lote_dict["itens_coletados"] = itens_coletados
        lotes.append(lote_dict)

    conn.close()
    return render_template("picking.html", lote=None, lotes=lotes)


@app.route("/picking/<numero_lote>")
def picking_lote(numero_lote):
    numero_lote = str(numero_lote or "").strip()

    conn = sqlite3.connect("status.db")
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()

    cursor.execute("SELECT * FROM lotes_envio WHERE numero_lote = ?", (numero_lote,))
    lote_row = cursor.fetchone()

    if not lote_row:
        conn.close()
        return redirect("/picking")

    lote = dict(lote_row)
    etapa_atual = lote.get("etapa_atual") or TIMELINE_ETAPAS[0]
    lote["etapa_atual"] = etapa_atual
    lote["status"] = obter_status_abertura_por_etapa(etapa_atual)

    cursor.execute("SELECT COUNT(*) FROM lotes_picking_itens WHERE numero_lote = ?", (numero_lote,))
    total_itens = int(cursor.fetchone()[0] or 0)

    if total_itens == 0:
        itens_snapshot = carregar_itens_snapshot_lote(numero_lote)
        if itens_snapshot:
            df_boot = pd.DataFrame(itens_snapshot)
            sincronizar_picking_itens(numero_lote, df_boot)

    cursor.execute("""
        SELECT numero_lote, codigo, sku, endereco, titulo, conta, selo, quantidade, observacao, coletado, coletado_em, quantidade_informada, divergencia, divergencia_em
        FROM lotes_picking_itens
        WHERE numero_lote = ?
        ORDER BY coletado ASC, endereco ASC, sku ASC, codigo ASC
    """, (numero_lote,))
    itens = [dict(row) for row in cursor.fetchall()]

    conn.close()
    return render_template("picking.html", lote=lote, itens=itens, lotes=[])


@app.route("/api/picking/coletar", methods=["POST"])
def api_picking_coletar():
    data = request.get_json() or {}
    numero_lote = str(data.get("numero_lote", "")).strip()
    sku_digitado = str(data.get("sku", "") or "").strip()
    observacao = str(data.get("observacao", "") or "").strip()

    try:
        quantidade_informada = int(float(str(data.get("quantidade", 0) or 0).replace(",", ".")))
    except:
        quantidade_informada = 0

    if not numero_lote or not sku_digitado:
        return jsonify({"ok": False, "erro": "Informe o lote e o SKU."}), 400

    if quantidade_informada <= 0:
        return jsonify({"ok": False, "erro": "Informe uma quantidade válida para a coleta."}), 400

    conn = sqlite3.connect("status.db")
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    garantir_colunas_gestao_operacional(cursor)

    usuario_atual = get_current_user()
    nome_coletor = ""
    try:
        nome_coletor = str(getattr(usuario_atual, "nome", "") or getattr(usuario_atual, "email", "") or "").strip()
    except:
        nome_coletor = ""

    cursor.execute(
        """
        SELECT *
        FROM lotes_picking_itens
        WHERE numero_lote = ?
          AND UPPER(TRIM(COALESCE(sku, ''))) = UPPER(TRIM(?))
        ORDER BY coletado ASC, divergencia ASC, codigo ASC
        """,
        (numero_lote, sku_digitado)
    )
    encontrados = cursor.fetchall()

    item = None
    for row in encontrados:
        if int(row["coletado"] or 0) != 1 and int(row["divergencia"] or 0) != 1:
            item = row
            break

    if not item:
        conn.close()
        return jsonify({"ok": False, "erro": "SKU não encontrado entre os itens pendentes deste lote."}), 404

    quantidade_esperada = int(item["quantidade"] or 0)
    agora = agora_str()

    if quantidade_informada > quantidade_esperada:
        conn.close()
        return jsonify({"ok": False, "erro": f"A quantidade informada ({quantidade_informada}) é maior que a esperada para este SKU ({quantidade_esperada})."}), 400

    divergencia_registrada = False
    mensagem = ""

    if quantidade_informada < quantidade_esperada:
        if not observacao:
            conn.close()
            return jsonify({
                "ok": False,
                "erro": f"Quantidade informada menor que a esperada. Esperado: {quantidade_esperada}. Informe na observação o motivo da divergência para registrar a ocorrência."
            }), 400

        observacao_final = f"DIVERGÊNCIA DE PICKING | Esperado: {quantidade_esperada} | Informado: {quantidade_informada} | Motivo: {observacao}"
        cursor.execute(
            """
            UPDATE lotes_picking_itens
            SET quantidade_informada = ?,
                divergencia = 1,
                divergencia_em = ?,
                divergencia_por = ?,
                observacao = ?
            WHERE numero_lote = ? AND codigo = ?
            """,
            (quantidade_informada, agora, nome_coletor, observacao_final, numero_lote, item["codigo"])
        )
        divergencia_registrada = True
        mensagem = "Divergência registrada com sucesso. O item ficou salvo com observação para conferência."
    else:
        cursor.execute(
            """
            UPDATE lotes_picking_itens
            SET coletado = 1,
                coletado_em = ?,
                coletado_por = ?,
                observacao = ?,
                quantidade_informada = ?,
                divergencia = 0,
                divergencia_em = '',
                divergencia_por = ''
            WHERE numero_lote = ? AND codigo = ?
            """,
            (agora, nome_coletor, observacao, quantidade_informada, numero_lote, item["codigo"])
        )
        mensagem = "SKU coletado com sucesso."

    mover_item_picking_para_conferencia(cursor, numero_lote, item["codigo"])

    conn.commit()

    cursor.execute("SELECT COUNT(*) FROM lotes_picking_itens WHERE numero_lote = ?", (numero_lote,))
    total = int(cursor.fetchone()[0] or 0)

    cursor.execute("SELECT COUNT(*) FROM lotes_picking_itens WHERE numero_lote = ? AND coletado = 1", (numero_lote,))
    coletados = int(cursor.fetchone()[0] or 0)

    cursor.execute("SELECT COUNT(*) FROM lotes_picking_itens WHERE numero_lote = ? AND divergencia = 1", (numero_lote,))
    divergencias = int(cursor.fetchone()[0] or 0)

    cursor.execute("SELECT etapa_atual FROM lotes_envio WHERE numero_lote = ?", (numero_lote,))
    row_lote = cursor.fetchone()
    etapa_atual = (row_lote["etapa_atual"] if row_lote else TIMELINE_ETAPAS[0]) or TIMELINE_ETAPAS[0]

    conn.close()

    finalizado = total > 0 and (coletados + divergencias) >= total
    if finalizado:
        atualizar_etapa_lote(numero_lote, "EM CONFERÊNCIA")
        conn_sync = sqlite3.connect("status.db")
        conn_sync.row_factory = sqlite3.Row
        cursor_sync = conn_sync.cursor()
        garantir_lote_conferencia_e_itens(cursor_sync, numero_lote)
        conn_sync.commit()
        conn_sync.close()
        etapa_atual = "EM CONFERÊNCIA"

    return jsonify({
        "ok": True,
        "total": total,
        "coletados": coletados,
        "divergencias": divergencias,
        "finalizado": finalizado,
        "etapa_atual": etapa_atual,
        "divergencia": divergencia_registrada,
        "mensagem": mensagem,
        "codigo": str(item["codigo"] or "")
    })



@app.route("/embalagem")
def embalagem_lista():
    conn = sqlite3.connect("status.db")
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    garantir_tabela_embalagem(cursor)

    cursor.execute("""
        SELECT DISTINCT le.*
        FROM lotes_envio le
        INNER JOIN lotes_embalagem_itens ei ON ei.numero_lote = le.numero_lote
        WHERE COALESCE(ei.embalado, 0) = 0
        ORDER BY le.data_criacao DESC, le.numero_lote DESC
    """)
    lotes_rows = cursor.fetchall()

    lotes = []
    for lote in lotes_rows:
        lote_dict = dict(lote)
        numero_lote = lote_dict["numero_lote"]
        sincronizar_embalagem_itens(cursor, numero_lote)

        cursor.execute("SELECT COUNT(*) FROM lotes_embalagem_itens WHERE numero_lote = ?", (numero_lote,))
        total_itens = int(cursor.fetchone()[0] or 0)

        cursor.execute("SELECT COUNT(*) FROM lotes_embalagem_itens WHERE numero_lote = ? AND embalado = 1", (numero_lote,))
        itens_embalados = int(cursor.fetchone()[0] or 0)

        cursor.execute("SELECT COUNT(*) FROM lotes_embalagem_itens WHERE numero_lote = ? AND divergencia = 1", (numero_lote,))
        itens_divergencia = int(cursor.fetchone()[0] or 0)

        lote_dict["total_itens"] = total_itens
        lote_dict["itens_embalados"] = itens_embalados
        lote_dict["itens_divergencia"] = itens_divergencia
        lote_dict["itens_pendentes"] = max(total_itens - itens_embalados - itens_divergencia, 0)
        lotes.append(lote_dict)

    conn.commit()
    conn.close()
    return render_template("embalagem.html", lote=None, lotes=lotes, itens=[])


@app.route("/embalagem/<numero_lote>")
def embalagem_lote(numero_lote):
    numero_lote = str(numero_lote or "").strip()

    conn = sqlite3.connect("status.db")
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    garantir_tabela_embalagem(cursor)

    cursor.execute("SELECT * FROM lotes_envio WHERE numero_lote = ?", (numero_lote,))
    lote_row = cursor.fetchone()

    if not lote_row:
        conn.close()
        return redirect("/embalagem")

    sincronizar_embalagem_itens(cursor, numero_lote)
    conn.commit()

    lote = dict(lote_row)
    lote["status"] = obter_status_abertura_por_etapa(lote.get("etapa_atual") or "EMBALAGEM")
    lote["etapa_atual"] = lote.get("etapa_atual") or "EMBALAGEM"

    cursor.execute("""
        SELECT *
        FROM lotes_embalagem_itens
        WHERE numero_lote = ?
        ORDER BY embalado ASC, endereco ASC, sku ASC, codigo ASC
    """, (numero_lote,))
    itens = [dict(row) for row in cursor.fetchall()]

    conn.close()
    return render_template("embalagem.html", lote=lote, lotes=[], itens=itens)


@app.route("/api/embalagem/confirmar", methods=["POST"])
def api_embalagem_confirmar():
    data = request.get_json() or {}
    numero_lote = str(data.get("numero_lote", "")).strip()
    sku_digitado = str(data.get("sku", "") or "").strip()
    observacao = str(data.get("observacao", "") or "").strip()

    try:
        quantidade_embalada = int(float(str(data.get("quantidade", 0) or 0).replace(",", ".")))
    except:
        quantidade_embalada = 0

    if not numero_lote or not sku_digitado:
        return jsonify({"ok": False, "erro": "Informe o lote e o SKU."}), 400

    if quantidade_embalada <= 0:
        return jsonify({"ok": False, "erro": "Informe uma quantidade válida para a embalagem."}), 400

    conn = sqlite3.connect("status.db")
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    garantir_tabela_embalagem(cursor)
    sincronizar_embalagem_itens(cursor, numero_lote)

    cursor.execute("""
        SELECT *
        FROM lotes_embalagem_itens
        WHERE numero_lote = ?
          AND UPPER(TRIM(COALESCE(sku, ''))) = UPPER(TRIM(?))
        ORDER BY embalado ASC, divergencia ASC, codigo ASC
    """, (numero_lote, sku_digitado))
    encontrados = cursor.fetchall()

    item = None
    for row in encontrados:
        if int(row["embalado"] or 0) != 1 and int(row["divergencia"] or 0) != 1:
            item = row
            break

    if not item:
        conn.close()
        return jsonify({"ok": False, "erro": "SKU não encontrado entre os itens pendentes desta embalagem."}), 404

    quantidade_esperada = int(item["quantidade"] or 0)
    agora = agora_str()
    nome_usuario = obter_nome_usuario_atual()

    if quantidade_embalada > quantidade_esperada:
        conn.close()
        return jsonify({"ok": False, "erro": f"A quantidade informada ({quantidade_embalada}) é maior que a esperada para este SKU ({quantidade_esperada})."}), 400

    divergencia_registrada = False
    mensagem = ""

    if quantidade_embalada < quantidade_esperada:
        if not observacao:
            conn.close()
            return jsonify({
                "ok": False,
                "erro": f"Quantidade embalada menor que a esperada. Esperado: {quantidade_esperada}. Informe o motivo na observação."
            }), 400

        observacao_final = f"DIVERGÊNCIA DE EMBALAGEM | Esperado: {quantidade_esperada} | Informado: {quantidade_embalada} | Motivo: {observacao}"
        cursor.execute("""
            UPDATE lotes_embalagem_itens
            SET quantidade_embalada = ?,
                divergencia = 1,
                divergencia_em = ?,
                observacao = ?
            WHERE numero_lote = ? AND codigo = ?
        """, (quantidade_embalada, agora, observacao_final, numero_lote, item["codigo"]))
        divergencia_registrada = True
        mensagem = "Divergência registrada na embalagem."
    else:
        cursor.execute("""
            UPDATE lotes_embalagem_itens
            SET embalado = 1,
                embalado_em = ?,
                embalado_por = ?,
                quantidade_embalada = ?,
                observacao = ?,
                divergencia = 0,
                divergencia_em = ''
            WHERE numero_lote = ? AND codigo = ?
        """, (agora, nome_usuario, quantidade_embalada, observacao, numero_lote, item["codigo"]))
        mensagem = "SKU embalado com sucesso."

    conn.commit()

    cursor.execute("SELECT COUNT(*) FROM lotes_embalagem_itens WHERE numero_lote = ?", (numero_lote,))
    total = int(cursor.fetchone()[0] or 0)

    cursor.execute("SELECT COUNT(*) FROM lotes_embalagem_itens WHERE numero_lote = ? AND embalado = 1", (numero_lote,))
    embalados = int(cursor.fetchone()[0] or 0)

    cursor.execute("SELECT COUNT(*) FROM lotes_embalagem_itens WHERE numero_lote = ? AND divergencia = 1", (numero_lote,))
    divergencias = int(cursor.fetchone()[0] or 0)

    conn.close()

    finalizado = total > 0 and (embalados + divergencias) >= total
    etapa_atual = "EMBALAGEM"
    if finalizado:
        atualizar_etapa_lote(numero_lote, "CONFERIR CAIXAS MASTER")
        etapa_atual = "CONFERIR CAIXAS MASTER"

    return jsonify({
        "ok": True,
        "total": total,
        "embalados": embalados,
        "divergencias": divergencias,
        "finalizado": finalizado,
        "etapa_atual": etapa_atual,
        "divergencia": divergencia_registrada,
        "mensagem": mensagem,
        "codigo": str(item["codigo"] or "")
    })



@app.route("/dados")
def dados():
    try:
        df = carregar_dados_base()
        df = df.fillna("")
        return jsonify(df.to_dict(orient="records"))
    except Exception as e:
        return jsonify({"ok": False, "erro": f"Erro ao carregar dados: {str(e)}", "dados": []}), 500


@app.route("/dados-dashboard")
def dados_dashboard():
    df = carregar_dados_base()
    return jsonify(df.to_dict(orient="records"))

@app.route("/api/coletas-calendario")
def api_coletas_calendario():
    conn = sqlite3.connect("status.db")
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()

    try:
        cursor.execute("""
            SELECT
                numero_lote,
                COALESCE(total_mlbs, 0) AS total_mlbs,
                COALESCE(total_pecas, 0) AS total_pecas,
                COALESCE(data_coleta_agendada, '') AS data_coleta_agendada,
                COALESCE(observacao, '') AS observacao,
                COALESCE(status_expedicao, '') AS status_expedicao,
                COALESCE(etapa_atual, '') AS etapa_atual,
                COALESCE(status, '') AS status
            FROM lotes_envio
            WHERE COALESCE(TRIM(data_coleta_agendada), '') <> ''
            ORDER BY data_coleta_agendada, numero_lote
        """)
        rows = cursor.fetchall()
    finally:
        conn.close()

    eventos = []
    for row in rows:
        row = dict(row)
        status_texto = normalizar_texto(row.get("status_expedicao") or row.get("etapa_atual") or row.get("status"))
        status_coleta = "COLETADO" if "COLETADO" in status_texto else "NAO_COLETADO"
        eventos.append({
            "numero_lote": str(row.get("numero_lote") or "").strip(),
            "total_mlbs": int(row.get("total_mlbs") or 0),
            "total_pecas": int(row.get("total_pecas") or 0),
            "data_coleta_agendada": str(row.get("data_coleta_agendada") or "").strip(),
            "observacao": str(row.get("observacao") or "").strip(),
            "status_coleta": status_coleta
        })

    return jsonify(eventos)


@app.route("/api/coletas-calendario/salvar", methods=["POST"])
def api_salvar_coleta_calendario():
    data = request.json if request.is_json else request.form

    numero_lote = str(data.get("numero_lote", "") or "").strip()
    data_coleta_agendada = str(data.get("data_coleta_agendada", "") or "").strip()
    total_mlbs = int(float(data.get("total_mlbs", 0) or 0))
    total_pecas = int(float(data.get("total_pecas", 0) or 0))
    observacao = str(data.get("observacao", "") or "").strip()
    status_coleta = normalizar_texto(data.get("status_coleta", "NAO_COLETADO"))

    if not numero_lote:
        return jsonify({"ok": False, "erro": "Número do lote é obrigatório."}), 400

    if not parse_data_hora(data_coleta_agendada):
        return jsonify({"ok": False, "erro": "Data da coleta inválida."}), 400

    etapa_destino = "COLETADO" if status_coleta == "COLETADO" else "AGUARDANDO COLETA"
    agora = agora_str()

    conn = sqlite3.connect("status.db")
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()

    cursor.execute("SELECT * FROM lotes_envio WHERE numero_lote = ?", (numero_lote,))
    existente = cursor.fetchone()

    if existente:
        timeline = carregar_timeline_json(existente["timeline_json"])
        timeline = sincronizar_timeline_ate_etapa(timeline, etapa_destino, agora)
        cursor.execute(
            """
            UPDATE lotes_envio
            SET total_mlbs = ?,
                total_pecas = ?,
                observacao = ?,
                data_coleta_agendada = ?,
                etapa_atual = ?,
                timeline_json = ?
            WHERE numero_lote = ?
            """,
            (
                total_mlbs,
                total_pecas,
                observacao,
                data_coleta_agendada,
                etapa_destino,
                timeline_para_json(timeline),
                numero_lote
            )
        )
    else:
        timeline = timeline_vazio()
        timeline = sincronizar_timeline_ate_etapa(timeline, etapa_destino, agora)
        cursor.execute(
            """
            INSERT INTO lotes_envio (
                numero_lote, tipo_lote, total_mlbs, total_pecas,
                status, responsavel, transportadora, observacao,
                prioridade, data_envio, status_expedicao, status_ecommerce,
                origem, data_criacao, timeline_json, etapa_atual, data_coleta_agendada
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                numero_lote,
                "Diversos",
                total_mlbs,
                total_pecas,
                obter_status_abertura_por_etapa(etapa_destino),
                "",
                "",
                observacao,
                "",
                "",
                "AGUARDANDO",
                "AGUARDANDO",
                "MANUAL",
                agora,
                timeline_para_json(timeline),
                etapa_destino,
                data_coleta_agendada
            )
        )

    atualizar_statuss_por_etapa(cursor, numero_lote, etapa_destino)
    conn.commit()
    conn.close()

    return jsonify({"ok": True, "numero_lote": numero_lote})


@app.route("/exportar-excel")
def exportar_excel():
    df = carregar_dados_base()

    conn = sqlite3.connect("status.db")
    cursor = conn.cursor()

    try:
        cursor.execute("SELECT codigo, status, quantidade, estrategia, motivo_envio, prioridade FROM status_cards")
        rows = cursor.fetchall()

        mapa_quantidade = {str(row[0]): row[2] or 0 for row in rows}
        mapa_status = {str(row[0]): row[1] or "" for row in rows}
        mapa_estrategia = {str(row[0]): row[3] or "" for row in rows}
        mapa_motivo_envio = {str(row[0]): row[4] or "" for row in rows}
    except:
        mapa_quantidade = {}
        mapa_status = {}
        mapa_estrategia = {}
        mapa_motivo_envio = {}

    try:
        mapa_lotes_exportacao = montar_mapa_lotes_exportacao(cursor)
    except:
        mapa_lotes_exportacao = {}

    try:
        cursor.execute("SELECT sku, comentario FROM comentarios")
        rows_comentarios = cursor.fetchall()
        mapa_comentarios = {str(sku): comentario or "" for sku, comentario in rows_comentarios}
    except:
        mapa_comentarios = {}

    try:
        cursor.execute("SELECT codigo, comentario FROM comentarios_mlb")
        rows_comentarios_mlb = cursor.fetchall()
        mapa_comentarios_mlb = {str(codigo): comentario or "" for codigo, comentario in rows_comentarios_mlb}
    except:
        mapa_comentarios_mlb = {}

    conn.close()

    if "Código do Anúncio" not in df.columns:
        df["Código do Anúncio"] = ""

    df["Quantidade para Enviar"] = pd.to_numeric(
        df["Código do Anúncio"].astype(str).map(mapa_quantidade), errors="coerce"
    ).fillna(0).astype(int)
    df["Estratégia"] = df["Código do Anúncio"].astype(str).map(mapa_estrategia).fillna("")
    df["Motivo do Envio"] = df["Código do Anúncio"].astype(str).map(mapa_motivo_envio).fillna("")
    df["Motivo do Envio"] = df["Motivo do Envio"].where(df["Motivo do Envio"].astype(str).str.strip() != "", df["Estratégia"])
    df["Motivo"] = df["Motivo do Envio"]

    tela = request.args.get("tela", "").strip()
    conta = request.args.get("conta", "").strip()
    busca = request.args.get("busca", "").strip().lower()
    selo = request.args.get("selo", "").strip()
    logica = request.args.get("logica", "").strip()
    saude = request.args.get("saude", "").strip()
    condicao = request.args.get("condicao", "").strip()
    ultimos = request.args.get("ultimos", "").strip()
    lote = request.args.get("lote", "").strip()
    magic = request.args.get("magic", "").strip()
    cobertura = request.args.get("cobertura", "").strip()
    full = request.args.get("full", "").strip()
    saude_full = request.args.get("saudeFull", "").strip()
    status_full_filtro = request.args.get("statusFullFiltro", "").strip()
    filtro_especial = request.args.get("filtroEspecial", "").strip()

    def valor_para_numero_excel(valor):
        if valor is None or valor == "":
            return 0
        try:
            return float(str(valor).replace(".", "").replace(",", "."))
        except:
            return 0

    def destino(codigo):
        status = mapa_status.get(str(codigo), "")
        if status == "enviando":
            return "enviando"
        if status == "ecommerce":
            return "ecommerce"
        if status == "compras":
            return "compras"
        if status == "acompanhamento":
            return "acompanhamento"
        if status == "homologar":
            return "homologar"
        if status in ["nao_enviar", "naoEnviar"]:
            return "naoEnviar"
        if status == "filetado":
            return "historico"
        return "principal"

    # Para a exportação completa em múltiplas abas, o filtro de tela não é aplicado aqui.
    # Assim, o mesmo arquivo leva todas as telas, cada uma em sua própria aba,
    # mantendo apenas os demais filtros compartilhados.

    if conta:
        df = df[df["Nickname"].astype(str).str.strip().str.upper() == conta.upper()]

    if busca:
        mask_busca = (
            df["SKU"].astype(str).str.lower().str.contains(busca, na=False) |
            df["Código do Anúncio"].astype(str).str.lower().str.contains(busca, na=False) |
            df["Título"].astype(str).str.lower().str.contains(busca, na=False)
        )
        df = df[mask_busca]

    if selo:
        df = df[df["SELO"].astype(str).str.strip() == selo]

    if logica:
        df = df[df["ANALISE"].astype(str).str.strip() == logica]

    if saude:
        df = df[df["SAUDE DO ESTOQUE 4i"].astype(str).str.strip() == saude]

    if condicao:
        coluna_condicao = "CONDIÇÃO" if "CONDIÇÃO" in df.columns else "CONDIÇAO"
        if coluna_condicao not in df.columns:
            df[coluna_condicao] = ""
        df = df[df[coluna_condicao].astype(str).str.strip() == condicao]

    if ultimos:
        if "ULTIMOS" not in df.columns:
            df["ULTIMOS"] = ""
        df = df[df["ULTIMOS"].astype(str).str.strip() == ultimos]

    if lote:
        if "LOTE" not in df.columns:
            df["LOTE"] = ""
        df = df[df["LOTE"].astype(str).str.strip() == lote]

    if magic:
        if "MAGIC" not in df.columns:
            df["MAGIC"] = ""
        df = df[df["MAGIC"].astype(str).str.strip().str.upper() == magic.upper()]

    if full:
        if "Full" not in df.columns:
            df["Full"] = ""
        df = df[df["Full"].astype(str).str.strip().str.upper() == full.upper()]

    if saude_full:
        if "SAUDE_ESTOQUE_FULL" not in df.columns:
            df["SAUDE_ESTOQUE_FULL"] = ""
        df = df[df["SAUDE_ESTOQUE_FULL"].astype(str).str.strip().str.upper() == saude_full.upper()]

    if status_full_filtro == "nao_ofereco_mais_full":
        if "OBSERVAÇÃO MELI" not in df.columns:
            df["OBSERVAÇÃO MELI"] = ""
        df = df[df["OBSERVAÇÃO MELI"].astype(str).str.strip() == "Você deixou de oferecer o Full."]

    if status_full_filtro == "esta_no_full":
        if "Full" not in df.columns:
            df["Full"] = ""
        df = df[df["Full"].astype(str).str.strip().str.upper() == "NO FULL"]

    if status_full_filtro == "nunca_foi_full":
        if "OBSERVAÇÃO MELI" not in df.columns:
            df["OBSERVAÇÃO MELI"] = ""
        if "Full" not in df.columns:
            df["Full"] = ""
        df = df[
            (df["OBSERVAÇÃO MELI"].astype(str).str.strip() != "Você deixou de oferecer o Full.") &
            (df["Full"].astype(str).str.strip().str.upper() != "NO FULL")
        ]

    if cobertura == "baixo30":
        coluna_cobertura = df["Cobertura"].astype(str) if "Cobertura" in df.columns else pd.Series([""] * len(df), index=df.index)
        dias = coluna_cobertura.str.extract(r"(\d+)\s*dias", expand=False)
        dias = pd.to_numeric(dias, errors="coerce")
        df = df[(dias.notna()) & (dias < 30)]

    if filtro_especial:
        unidades = pd.to_numeric(df["30 DIAS"], errors="coerce").fillna(0) if "30 DIAS" in df.columns else pd.Series([0] * len(df), index=df.index)
        valores = df["Total de Vendas 30 DIAS"].apply(valor_para_numero_excel) if "Total de Vendas 30 DIAS" in df.columns else pd.Series([0] * len(df), index=df.index)

        if filtro_especial == "valor500":
            df = df[valores > 500]
        elif filtro_especial == "unidades10":
            df = df[unidades > 10]
        elif filtro_especial == "critico":
            df = df[(unidades < 10) & (valores > 500)]

    output = BytesIO()

    abas = [
        ("principal", "Principal"),
        ("enviando", "Enviando"),
        ("ecommerce", "Ecommerce"),
        ("compras", "Compras"),
        ("acompanhamento", "Acompanhamento"),
        ("homologar", "Homologar"),
        ("naoEnviar", "Nao Enviar")
    ]

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        for chave_tela, nome_aba in abas:
            df_aba = df[df["Código do Anúncio"].apply(destino) == chave_tela].copy()
            df_export = preparar_dataframe_exportacao(df_aba, mapa_lotes_exportacao)
            df_export.to_excel(writer, index=False, sheet_name=nome_aba)

            worksheet = writer.sheets[nome_aba]
            for idx, coluna in enumerate(df_export.columns, start=1):
                largura = max(len(str(coluna)), 14)
                if not df_export.empty:
                    maior_valor = df_export[coluna].astype(str).map(len).max()
                    if pd.notna(maior_valor):
                        largura = max(largura, min(int(maior_valor) + 2, 40))
                worksheet.column_dimensions[chr(64 + idx)].width = largura

    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name="relatorio_envio.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


def registrar_lote_conferencia(numero_lote, tipo_lote, df_lote):
    if df_lote.empty:
        return

    conn = sqlite3.connect("status.db")
    cursor = conn.cursor()

    agora = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    cursor.execute("""
        INSERT INTO lotes_conferencia (numero_lote, tipo_lote, status, data_criacao)
        VALUES (?, ?, ?, ?)
        ON CONFLICT(numero_lote) DO UPDATE SET
            tipo_lote=excluded.tipo_lote
    """, (numero_lote, tipo_lote, "PENDENTE", agora))

    for _, item in df_lote.iterrows():
        codigo = str(item.get("Código do Anúncio", "") or "")
        sku = str(item.get("SKU", "") or "")
        titulo = str(item.get("Título", "") or "")
        endereco = str(item.get("ENDEREÇO", "") or "")
        lote_filete = str(item.get("Lote", "") or "")
        quantidade_esperada = int(float(item.get("Enviar", 0) or 0))

        cursor.execute("""
            INSERT INTO lotes_itens (
                numero_lote, codigo, sku, titulo,
                quantidade_esperada, endereco, lote_filete
            )
            VALUES (?, ?, ?, ?, ?, ?, ?)
            ON CONFLICT(numero_lote, codigo) DO UPDATE SET
                sku=excluded.sku,
                titulo=excluded.titulo,
                quantidade_esperada=excluded.quantidade_esperada,
                endereco=excluded.endereco,
                lote_filete=excluded.lote_filete
        """, (
            numero_lote, codigo, sku, titulo,
            quantidade_esperada, endereco, lote_filete
        ))

    conn.commit()
    conn.close()



def atualizar_lote_envio_existente(numero_lote, tipo_lote, df_lote):
    if df_lote.empty:
        return

    conn = sqlite3.connect("status.db")
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()

    cursor.execute("SELECT numero_lote, timeline_json, etapa_atual, data_coleta_agendada FROM lotes_envio WHERE numero_lote = ?", (numero_lote,))
    existe = cursor.fetchone()

    total_mlbs = int(len(df_lote))
    total_pecas = int(pd.to_numeric(df_lote["Enviar"], errors="coerce").fillna(0).sum())
    agora = agora_str()
    etapa_inicial = "GERAR LOTE MELI"

    if existe:
        timeline = carregar_timeline_json(existe["timeline_json"])
        etapa_atual = existe["etapa_atual"] or etapa_inicial
        timeline = sincronizar_timeline_ate_etapa(timeline, etapa_atual, agora)

        cursor.execute(
            """
            UPDATE lotes_envio
            SET tipo_lote = ?,
                total_mlbs = ?,
                total_pecas = ?,
                origem = 'FILETE',
                etapa_atual = ?,
                timeline_json = ?
            WHERE numero_lote = ?
            """,
            (tipo_lote, total_mlbs, total_pecas, etapa_atual, timeline_para_json(timeline), numero_lote)
        )
        atualizar_statuss_por_etapa(cursor, numero_lote, etapa_atual)
    else:
        timeline = timeline_vazio()
        timeline = sincronizar_timeline_ate_etapa(timeline, etapa_inicial, agora)

        cursor.execute(
            """
            INSERT INTO lotes_envio (
                numero_lote, tipo_lote, total_mlbs, total_pecas,
                status, responsavel, transportadora, observacao,
                prioridade, data_envio, status_expedicao,
                status_ecommerce, origem, data_criacao,
                timeline_json, etapa_atual, data_coleta_agendada
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                numero_lote,
                tipo_lote,
                total_mlbs,
                total_pecas,
                obter_status_abertura_por_etapa(etapa_inicial),
                "",
                "",
                "Gerado via filete",
                "",
                "",
                "AGUARDANDO",
                etapa_inicial,
                "FILETE",
                agora,
                timeline_para_json(timeline),
                etapa_inicial,
                ""
            )
        )

    conn.commit()
    conn.close()


def salvar_historico_e_finalizar_envio(numero_lote, tipo_lote, df_lote):
    if df_lote.empty:
        return

    conn = sqlite3.connect("status.db")
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()

    agora = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    cursor.execute("""
        SELECT codigo, estrategia, motivo_envio
        FROM status_cards
    """)
    rows_status = cursor.fetchall()
    mapa_status = {
        str(row["codigo"]): {
            "estrategia": row["estrategia"] or "",
            "motivo_envio": row["motivo_envio"] or ""
        }
        for row in rows_status
    }

    cursor.execute("""
        SELECT codigo, comentario
        FROM comentarios_mlb
    """)
    rows_comentarios = cursor.fetchall()
    mapa_comentarios_mlb = {
        str(row["codigo"]): row["comentario"] or ""
        for row in rows_comentarios
    }

    cursor.execute("""
        DELETE FROM lotes_envio_itens_snapshot
        WHERE numero_lote = ?
    """, (numero_lote,))

    def valor_json(valor):
        try:
            if pd.isna(valor):
                return ""
        except:
            pass

        if isinstance(valor, (int, float, str, bool)) or valor is None:
            return valor

        return str(valor)

    for _, item in df_lote.iterrows():
        codigo = str(item.get("Código do Anúncio", "") or "")
        sku = str(item.get("SKU", "") or "")
        titulo = str(item.get("Título", "") or "")
        nickname = str(item.get("Nickname", "") or "")
        endereco = str(item.get("ENDEREÇO", "") or "")
        lote_filete = str(item.get("Lote", "") or "")
        quantidade = int(float(item.get("Enviar", 0) or 0))

        estrategia = mapa_status.get(codigo, {}).get("estrategia", "")
        motivo_envio = mapa_status.get(codigo, {}).get("motivo_envio", "") or estrategia
        comentario_mlb = mapa_comentarios_mlb.get(codigo, "")

        dados_item = {coluna: valor_json(valor) for coluna, valor in item.to_dict().items()}

        cursor.execute("""
            INSERT INTO lotes_envio_itens_snapshot (
                numero_lote, tipo_lote, codigo, sku, titulo,
                nickname, quantidade, endereco, lote_filete,
                estrategia, motivo_envio, comentario_mlb,
                dados_json, data_geracao
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            numero_lote,
            tipo_lote,
            codigo,
            sku,
            titulo,
            nickname,
            quantidade,
            endereco,
            lote_filete,
            estrategia,
            motivo_envio,
            comentario_mlb,
            json.dumps(dados_item, ensure_ascii=False),
            agora
        ))

        cursor.execute("""
            UPDATE status_cards
            SET status = ?, quantidade = ?
            WHERE codigo = ?
        """, ("principal", 0, codigo))

    conn.commit()
    conn.close()
    sincronizar_picking_itens(numero_lote, df_lote)

@app.route("/criar-lote-enviando", methods=["POST"])
def criar_lote_enviando():
    data = request.get_json() or {}

    tipo_lote_padrao = str(data.get("tipo_lote", "Diversos")).strip() or "Diversos"
    lotes_config = data.get("lotes", []) or []

    df = carregar_dados_base()

    conn = sqlite3.connect("status.db")
    cursor = conn.cursor()
    cursor.execute("SELECT codigo, status, quantidade, estrategia FROM status_cards")
    rows = cursor.fetchall()
    conn.close()

    mapa_status = {str(codigo): status or "" for codigo, status, quantidade, estrategia in rows}
    mapa_quantidade = {str(codigo): quantidade or 0 for codigo, status, quantidade, estrategia in rows}

    df["STATUS_CARD"] = df["Código do Anúncio"].astype(str).map(mapa_status).fillna("principal")
    df["Enviar"] = df["Código do Anúncio"].astype(str).map(mapa_quantidade).fillna(0)

    df = df[(df["STATUS_CARD"] == "enviando") & (df["Enviar"].astype(float) > 0)].copy()

    if df.empty:
        return jsonify({"ok": False, "erro": "Não há itens na tela Enviando para criar o lote."}), 400

    colunas_necessarias = [
        "Nickname",
        "Código do Anúncio",
        "SKU",
        "Título",
        "ENDEREÇO",
        "SELO"
    ]

    for col in colunas_necessarias:
        if col not in df.columns:
            df[col] = ""

    def classificar_tipo_item(row):
        selo = str(row.get("SELO", "") or "").strip().upper()
        if "CAIXA" in selo:
            return "Caixa"
        if "DIVERS" in selo:
            return "Diversos"
        return tipo_lote_padrao

    df["CONTA_LOTE"] = df["Nickname"].astype(str).str.strip()
    df["CONTA_LOTE"] = df["CONTA_LOTE"].replace("", "SEM CONTA")
    df["TIPO_LOTE_GRUPO"] = df.apply(classificar_tipo_item, axis=1)

    if not isinstance(lotes_config, list) or not lotes_config:
        return jsonify({"ok": False, "erro": "Informe os números dos lotes para cada conta e tipo."}), 400

    mapa_lotes = {}
    numeros_utilizados = set()

    for item in lotes_config:
        conta = str((item or {}).get("conta", "") or "").strip() or "SEM CONTA"
        tipo = str((item or {}).get("tipo", "") or "").strip() or tipo_lote_padrao
        numero_lote = str((item or {}).get("numero_lote", "") or "").strip()

        if not numero_lote:
            return jsonify({"ok": False, "erro": f"Informe o número do lote para {conta} / {tipo}."}), 400

        numero_normalizado = numero_lote.upper()
        if numero_normalizado in numeros_utilizados:
            return jsonify({"ok": False, "erro": f"O número de lote {numero_lote} foi informado mais de uma vez."}), 400

        numeros_utilizados.add(numero_normalizado)
        mapa_lotes[(conta.upper(), tipo.upper())] = numero_lote

    grupos = []
    for (conta, tipo), df_grupo in df.groupby(["CONTA_LOTE", "TIPO_LOTE_GRUPO"], sort=True):
        chave = (str(conta).upper(), str(tipo).upper())
        numero_lote = mapa_lotes.get(chave)

        if not numero_lote:
            return jsonify({"ok": False, "erro": f"Faltou informar o número do lote para {conta} / {tipo}."}), 400

        df_lote = df_grupo.copy()
        df_lote["Lote"] = f"Lote {tipo} - #{numero_lote}"

        if "ENDEREÇO" in df_lote.columns:
            df_lote = df_lote.sort_values(by="ENDEREÇO", kind="stable")

        try:
            registrar_lote_conferencia(numero_lote, tipo, df_lote)
            atualizar_lote_envio_existente(numero_lote, tipo, df_lote)
            salvar_historico_e_finalizar_envio(numero_lote, tipo, df_lote)
        except ValueError as e:
            return jsonify({"ok": False, "erro": str(e)}), 400

        grupos.append({
            "numero_lote": numero_lote,
            "conta": conta,
            "tipo": tipo,
            "itens": int(len(df_lote))
        })

    return jsonify({"ok": True, "lotes_criados": grupos})

def _valor_primeiro(row, nomes):
    for nome in nomes:
        if nome in row and str(row.get(nome, "") or "").strip() != "":
            return row.get(nome)
    return ""


def _formatar_data_filete(valor):
    texto = str(valor or "").strip()
    if not texto:
        return ""
    dt = parse_data_hora(texto)
    if dt:
        return dt.strftime("%d/%m/%Y")
    return texto


def _normalizar_conta_filete(conta):
    texto = str(conta or "").strip().upper()
    if not texto:
        return "SEM CONTA"
    return texto


def _copiar_template_filete():
    caminhos = [
        os.path.join(app.root_path, "PICKING.xlsx"),
        os.path.join(os.getcwd(), "PICKING.xlsx"),
        os.path.join(os.path.dirname(__file__), "PICKING.xlsx"),
        os.path.join("static", "PICKING.xlsx"),
    ]
    for caminho in caminhos:
        if caminho and os.path.exists(caminho):
            return load_workbook(caminho)
    raise FileNotFoundError("Arquivo modelo PICKING.xlsx não encontrado na pasta do sistema.")


def _linhas_inicio_blocos_filete(ws):
    inicios = []
    for row in range(1, ws.max_row + 1):
        valor = str(ws.cell(row=row, column=4).value or "").strip().upper()
        if valor == "ETAPA 2 - PICKING":
            inicios.append(row)
    if not inicios:
        inicios = list(range(1, ws.max_row + 1, 9))
    return inicios


def _limpar_blocos_filete(ws):
    for inicio in _linhas_inicio_blocos_filete(ws):
        for celula in [
            (inicio + 1, 5),
            (inicio + 2, 5),
            (inicio + 2, 7),
            (inicio + 3, 5),
            (inicio + 4, 5),
            (inicio + 4, 7),
            (inicio + 5, 5),
            (inicio + 5, 7),
        ]:
            ws.cell(*celula).value = ""


def _preencher_blocos_filete(ws, itens, conta_label, numero_lote, data_lote):
    inicios = _linhas_inicio_blocos_filete(ws)
    _limpar_blocos_filete(ws)

    for idx, (_, row) in enumerate(itens.iterrows()):
        if idx >= len(inicios):
            break
        inicio = inicios[idx]
        codigo = str(row.get("Código do Anúncio", "") or "").strip()
        sku = row.get("SKU", "")
        titulo = str(row.get("Título", "") or "").strip()
        endereco = str(row.get("ENDEREÇO", "") or "").strip()
        try:
            enviar = int(float(str(row.get("Enviar", 0) or 0).replace(",", ".")))
        except:
            enviar = row.get("Enviar", "") or ""

        ws.cell(inicio + 1, 5).value = conta_label
        ws.cell(inicio + 2, 5).value = numero_lote
        ws.cell(inicio + 2, 7).value = _formatar_data_filete(data_lote)
        ws.cell(inicio + 3, 5).value = titulo
        ws.cell(inicio + 4, 5).value = codigo
        ws.cell(inicio + 4, 7).value = sku
        ws.cell(inicio + 5, 5).value = endereco
        ws.cell(inicio + 5, 7).value = enviar


def _preencher_base_filete(wb, df, numero_lote, tipo_lote, data_lote, conta_prefixo=""):
    if "BASE" not in wb.sheetnames:
        return
    ws = wb["BASE"]
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)

    data_formatada = _formatar_data_filete(data_lote)
    for idx, (_, row) in enumerate(df.iterrows(), start=2):
        conta = str(row.get("Nickname", "") or "").strip()
        codigo = str(row.get("Código do Anúncio", "") or "").strip()
        sku = row.get("SKU", "")
        try:
            enviar = int(float(str(row.get("Enviar", 0) or 0).replace(",", ".")))
        except:
            enviar = row.get("Enviar", "") or ""
        endereco = str(row.get("ENDEREÇO", "") or "").strip()
        titulo = str(row.get("Título", "") or "").strip()

        valores = [
            conta,
            codigo,
            sku,
            enviar,
            row.get("ESTOQUE TOTAL SIGNUS", ""),
            row.get("APÓS ENVIO", ""),
            endereco,
            tipo_lote,
            titulo,
            row.get("Motivo", ""),
            row.get("RETORNO", ""),
            numero_lote,
            data_formatada,
            data_formatada,
            f"{conta_prefixo}{conta}".strip(),
            conta_prefixo,
        ]
        for col, valor in enumerate(valores, start=1):
            ws.cell(row=idx, column=col).value = valor


def montar_excel_filete_antigo(df, numero_lote, tipo_lote, data_lote=""):
    wb = _copiar_template_filete()

    for col in ["Nickname", "Código do Anúncio", "SKU", "Título", "ENDEREÇO", "Enviar"]:
        if col not in df.columns:
            df[col] = ""

    df = df.copy()
    df["CONTA_ARQ"] = df["Nickname"].apply(_normalizar_conta_filete)
    if "ENDEREÇO" in df.columns:
        df = df.sort_values(by=["CONTA_ARQ", "ENDEREÇO", "Código do Anúncio"], kind="stable")

    grupos = list(df.groupby("CONTA_ARQ", sort=True))
    if not grupos:
        grupos = [("SEM CONTA", df)]

    # Regra visual solicitada: a primeira conta usa o modelo azul (UNITARIO C1) e a segunda usa o modelo laranja (UNITARIO C2).
    abas_modelo = ["UNITARIO C1", "UNITARIO C2"]
    todas_abas_usadas = []
    for idx, (conta_key, df_conta) in enumerate(grupos):
        if idx < len(abas_modelo):
            aba = abas_modelo[idx]
            if aba not in wb.sheetnames:
                continue
            ws = wb[aba]
        else:
            origem = wb[abas_modelo[-1]]
            ws = wb.copy_worksheet(origem)
            ws.title = f"UNITARIO C{idx + 1}"
        todas_abas_usadas.append(ws.title)
        conta_original = str(df_conta.iloc[0].get("Nickname", "") or conta_key).strip() or conta_key
        conta_label = f"CONTA {idx + 1} - {conta_original}" if idx > 0 else conta_original
        _preencher_blocos_filete(ws, df_conta, conta_label, numero_lote, data_lote)

    for aba in abas_modelo:
        if aba in wb.sheetnames and aba not in todas_abas_usadas:
            _limpar_blocos_filete(wb[aba])

    _preencher_base_filete(wb, df, numero_lote, tipo_lote, data_lote)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


@app.route("/gerar-filete")
def gerar_filete():
    numero_lote = request.args.get("numero_lote", "").strip()
    tipo_lote = request.args.get("tipo_lote", "Diversos").strip() or "Diversos"
    data_lote = request.args.get("data_lote", "").strip()

    if not numero_lote:
        return "Informe um número de lote válido.", 400

    df = carregar_dados_base()

    conn = sqlite3.connect("status.db")
    cursor = conn.cursor()
    cursor.execute("SELECT codigo, status, quantidade, estrategia FROM status_cards")
    rows = cursor.fetchall()
    conn.close()

    mapa_status = {str(codigo): status or "" for codigo, status, quantidade, estrategia in rows}
    mapa_quantidade = {str(codigo): quantidade or 0 for codigo, status, quantidade, estrategia in rows}

    df["STATUS_CARD"] = df["Código do Anúncio"].astype(str).map(mapa_status).fillna("principal")
    df["Enviar"] = df["Código do Anúncio"].astype(str).map(mapa_quantidade).fillna(0)

    df = df[(df["STATUS_CARD"] == "enviando") & (df["Enviar"].astype(float) > 0)].copy()

    colunas_necessarias = [
        "Nickname",
        "Código do Anúncio",
        "SKU",
        "Título",
        "ENDEREÇO"
    ]

    for col in colunas_necessarias:
        if col not in df.columns:
            df[col] = ""

    df["Lote"] = f"Lote {tipo_lote} - #{numero_lote}" if numero_lote else f"Lote {tipo_lote}"

    if "ENDEREÇO" in df.columns:
        df = df.sort_values(by=["Nickname", "ENDEREÇO", "Código do Anúncio"], kind="stable")

    try:
        registrar_lote_conferencia(numero_lote, tipo_lote, df)
        atualizar_lote_envio_existente(numero_lote, tipo_lote, df)
        salvar_historico_e_finalizar_envio(numero_lote, tipo_lote, df)
    except ValueError as e:
        return str(e), 400

    output = montar_excel_filete_antigo(df, numero_lote, tipo_lote, data_lote)

    nome_arquivo = "filete.xlsx"
    if numero_lote:
        nome_arquivo = f"filete_{numero_lote}.xlsx"

    return send_file(
        output,
        as_attachment=True,
        download_name=nome_arquivo,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


def carregar_itens_snapshot_lote(numero_lote):
    conn = sqlite3.connect("status.db")
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()

    cursor.execute("""
        SELECT *
        FROM lotes_envio_itens_snapshot
        WHERE numero_lote = ?
        ORDER BY endereco, sku, codigo
    """, (numero_lote,))
    itens = [dict(row) for row in cursor.fetchall()]

    if not itens:
        cursor.execute("""
            SELECT
                li.numero_lote,
                COALESCE(le.tipo_lote, lc.tipo_lote, '') AS tipo_lote,
                li.codigo,
                li.sku,
                li.titulo,
                '' AS nickname,
                li.quantidade_esperada AS quantidade,
                li.endereco,
                li.lote_filete,
                '' AS estrategia,
                '' AS motivo_envio,
                '' AS comentario_mlb,
                '' AS dados_json,
                COALESCE(le.data_criacao, lc.data_criacao, '') AS data_geracao
            FROM lotes_itens li
            LEFT JOIN lotes_envio le
                ON le.numero_lote = li.numero_lote
            LEFT JOIN lotes_conferencia lc
                ON lc.numero_lote = li.numero_lote
            WHERE li.numero_lote = ?
            ORDER BY li.endereco, li.sku, li.codigo
        """, (numero_lote,))
        itens = [dict(row) for row in cursor.fetchall()]

    conn.close()

    for item in itens:
        item["dados_json"] = item.get("dados_json") or "{}"

    itens = enriquecer_itens_lote_com_base(itens)

    return itens
    return itens

def montar_excel_filete_lote(df, numero_lote, tipo_lote, data_lote=""):
    return montar_excel_filete_antigo(df, numero_lote, tipo_lote, data_lote)




def obter_dados_base_para_lote(codigo="", sku=""):
    codigo = str(codigo or "").strip()
    sku = str(sku or "").strip()

    try:
        df_base = carregar_dados_base().fillna("")
    except:
        return {}

    if codigo and "Código do Anúncio" in df_base.columns:
        df_codigo = df_base[df_base["Código do Anúncio"].astype(str).str.strip() == codigo]
        if not df_codigo.empty:
            return {str(col): df_codigo.iloc[0][col] for col in df_base.columns}

    if sku and "SKU" in df_base.columns:
        df_sku = df_base[df_base["SKU"].astype(str).str.strip() == sku]
        if not df_sku.empty:
            return {str(col): df_sku.iloc[0][col] for col in df_base.columns}

    return {}


def recalcular_totais_lote(cursor, numero_lote):
    cursor.execute("""
        SELECT COUNT(*), COALESCE(SUM(COALESCE(quantidade, 0)), 0)
        FROM lotes_envio_itens_snapshot
        WHERE numero_lote = ?
    """, (numero_lote,))
    row = cursor.fetchone()

    total_mlbs = int((row[0] if row else 0) or 0)
    total_pecas = int((row[1] if row else 0) or 0)

    cursor.execute("""
        UPDATE lotes_envio
        SET total_mlbs = ?, total_pecas = ?
        WHERE numero_lote = ?
    """, (total_mlbs, total_pecas, numero_lote))

    return total_mlbs, total_pecas


def montar_dataframe_snapshot_lote(numero_lote):
    itens = carregar_itens_snapshot_lote(numero_lote)
    registros = []

    for item in itens:
        dados_item = {}
        try:
            dados_item = json.loads(item.get("dados_json") or "{}")
            if not isinstance(dados_item, dict):
                dados_item = {}
        except:
            dados_item = {}

        dados_item["Código do Anúncio"] = item.get("codigo", "")
        dados_item["SKU"] = item.get("sku", "")
        dados_item["Título"] = item.get("titulo", "")
        dados_item["Nickname"] = item.get("nickname", "")
        dados_item["ENDEREÇO"] = item.get("endereco", "")
        dados_item["Lote"] = item.get("lote_filete", "")
        dados_item["Enviar"] = int(item.get("quantidade") or 0)
        registros.append(dados_item)

    return pd.DataFrame(registros)


@app.route("/api/lote-envio/<numero_lote>/item", methods=["POST"])
def api_adicionar_item_lote(numero_lote):
    data = request.get_json(silent=True) or {}

    numero_lote = str(numero_lote or "").strip()
    codigo = str(data.get("codigo", "") or "").strip()
    sku = str(data.get("sku", "") or "").strip()
    titulo_manual = str(data.get("titulo", "") or "").strip()
    nickname_manual = str(data.get("nickname", "") or "").strip()
    endereco_manual = str(data.get("endereco", "") or "").strip()
    estrategia_manual = str(data.get("estrategia", "") or "").strip()
    motivo_manual = str(data.get("motivo_envio", "") or "").strip()

    try:
        quantidade = int(float(str(data.get("quantidade", 0) or 0).replace(",", ".")))
    except:
        quantidade = 0

    if not numero_lote:
        return jsonify({"ok": False, "erro": "Número do lote não informado."}), 400

    if not codigo:
        return jsonify({"ok": False, "erro": "Informe o MLB."}), 400

    if quantidade <= 0:
        return jsonify({"ok": False, "erro": "Informe uma quantidade válida."}), 400

    conn = sqlite3.connect("status.db")
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()

    cursor.execute("SELECT numero_lote, tipo_lote FROM lotes_envio WHERE numero_lote = ?", (numero_lote,))
    lote = cursor.fetchone()
    if not lote:
        conn.close()
        return jsonify({"ok": False, "erro": "Lote não encontrado."}), 404

    cursor.execute("SELECT 1 FROM lotes_envio_itens_snapshot WHERE numero_lote = ? AND codigo = ?", (numero_lote, codigo))
    if cursor.fetchone():
        conn.close()
        return jsonify({"ok": False, "erro": "Este MLB já existe dentro deste lote."}), 400

    base_item = obter_dados_base_para_lote(codigo, sku)

    if not sku:
        sku = str(base_item.get("SKU", "") or "").strip()

    titulo = titulo_manual or str(base_item.get("Título", "") or "").strip()
    nickname = nickname_manual or str(base_item.get("Nickname", "") or "").strip()
    endereco = endereco_manual or str(base_item.get("ENDEREÇO", "") or "").strip()
    lote_filete = str(base_item.get("LOTE", "") or "").strip()

    cursor.execute("SELECT estrategia, motivo_envio FROM status_cards WHERE codigo = ?", (codigo,))
    row_status = cursor.fetchone()
    estrategia = estrategia_manual or (str(row_status["estrategia"] or "").strip() if row_status else "")
    motivo_envio = motivo_manual or (str(row_status["motivo_envio"] or "").strip() if row_status else "") or estrategia

    comentario_mlb = ""
    try:
        cursor.execute("SELECT comentario FROM comentarios_mlb WHERE codigo = ?", (codigo,))
        row_coment = cursor.fetchone()
        comentario_mlb = str((row_coment["comentario"] if row_coment else "") or "").strip()
    except:
        comentario_mlb = ""

    dados_item = {}
    for k, v in (base_item or {}).items():
        dados_item[str(k)] = v
    dados_item["Código do Anúncio"] = codigo
    dados_item["SKU"] = sku
    dados_item["Título"] = titulo
    dados_item["Nickname"] = nickname
    dados_item["ENDEREÇO"] = endereco
    dados_item["LOTE"] = lote_filete
    dados_item["Enviar"] = quantidade

    agora = agora_str()
    tipo_lote = str(lote["tipo_lote"] or "Diversos").strip() or "Diversos"

    cursor.execute("""
        INSERT INTO lotes_envio_itens_snapshot (
            numero_lote, tipo_lote, codigo, sku, titulo,
            nickname, quantidade, endereco, lote_filete,
            estrategia, motivo_envio, comentario_mlb,
            dados_json, data_geracao
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, (
        numero_lote, tipo_lote, codigo, sku, titulo,
        nickname, quantidade, endereco, lote_filete,
        estrategia, motivo_envio, comentario_mlb,
        json.dumps(dados_item, ensure_ascii=False), agora
    ))

    try:
        cursor.execute("""
            INSERT INTO lotes_itens (numero_lote, codigo, sku, titulo, quantidade_esperada, endereco, lote_filete)
            VALUES (?, ?, ?, ?, ?, ?, ?)
            ON CONFLICT(numero_lote, codigo) DO UPDATE SET
                sku = excluded.sku,
                titulo = excluded.titulo,
                quantidade_esperada = excluded.quantidade_esperada,
                endereco = excluded.endereco,
                lote_filete = excluded.lote_filete
        """, (numero_lote, codigo, sku, titulo, quantidade, endereco, lote_filete))
    except:
        pass

    try:
        cursor.execute("""
            INSERT INTO lotes_conferencia (numero_lote, tipo_lote, status, data_criacao)
            VALUES (?, ?, ?, ?)
            ON CONFLICT(numero_lote) DO UPDATE SET tipo_lote = excluded.tipo_lote
        """, (numero_lote, tipo_lote, "PENDENTE", agora))
    except:
        pass

    recalcular_totais_lote(cursor, numero_lote)

    conn.commit()
    conn.close()

    try:
        sincronizar_picking_itens(numero_lote, montar_dataframe_snapshot_lote(numero_lote))
    except:
        pass

    return jsonify({"ok": True})


@app.route("/api/lote-envio/<numero_lote>/item/<int:item_id>/quantidade", methods=["POST"])
def api_atualizar_quantidade_item_lote(numero_lote, item_id):
    data = request.get_json(silent=True) or {}

    try:
        quantidade = int(float(str(data.get("quantidade", 0) or 0).replace(",", ".")))
    except:
        quantidade = 0

    if quantidade <= 0:
        return jsonify({"ok": False, "erro": "Informe uma quantidade válida."}), 400

    conn = sqlite3.connect("status.db")
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()

    cursor.execute("SELECT * FROM lotes_envio_itens_snapshot WHERE id = ? AND numero_lote = ?", (item_id, numero_lote))
    item = cursor.fetchone()
    if not item:
        conn.close()
        return jsonify({"ok": False, "erro": "Item do lote não encontrado."}), 404

    dados_item = {}
    try:
        dados_item = json.loads(item["dados_json"] or "{}")
        if not isinstance(dados_item, dict):
            dados_item = {}
    except:
        dados_item = {}
    dados_item["Enviar"] = quantidade

    cursor.execute("""
        UPDATE lotes_envio_itens_snapshot
        SET quantidade = ?, dados_json = ?
        WHERE id = ? AND numero_lote = ?
    """, (quantidade, json.dumps(dados_item, ensure_ascii=False), item_id, numero_lote))

    try:
        cursor.execute("UPDATE lotes_itens SET quantidade_esperada = ? WHERE numero_lote = ? AND codigo = ?", (quantidade, numero_lote, item["codigo"]))
    except:
        pass

    recalcular_totais_lote(cursor, numero_lote)

    conn.commit()
    conn.close()

    try:
        sincronizar_picking_itens(numero_lote, montar_dataframe_snapshot_lote(numero_lote))
    except:
        pass

    return jsonify({"ok": True})


@app.route("/api/lote-envio/<numero_lote>/item/<int:item_id>", methods=["POST"])
def api_remover_item_lote(numero_lote, item_id):
    conn = sqlite3.connect("status.db")
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()

    cursor.execute("SELECT codigo FROM lotes_envio_itens_snapshot WHERE id = ? AND numero_lote = ?", (item_id, numero_lote))
    item = cursor.fetchone()
    if not item:
        conn.close()
        return jsonify({"ok": False, "erro": "Item do lote não encontrado."}), 404

    codigo = str(item["codigo"] or "").strip()

    cursor.execute("DELETE FROM lotes_envio_itens_snapshot WHERE id = ? AND numero_lote = ?", (item_id, numero_lote))
    try:
        cursor.execute("DELETE FROM lotes_itens WHERE numero_lote = ? AND codigo = ?", (numero_lote, codigo))
    except:
        pass
    try:
        cursor.execute("DELETE FROM lotes_picking_itens WHERE numero_lote = ? AND codigo = ?", (numero_lote, codigo))
    except:
        pass

    recalcular_totais_lote(cursor, numero_lote)

    conn.commit()
    conn.close()

    try:
        sincronizar_picking_itens(numero_lote, montar_dataframe_snapshot_lote(numero_lote))
    except:
        pass

    return jsonify({"ok": True})



@app.route("/lote-envio/<numero_lote>/pdf")
def lote_envio_pdf(numero_lote):
    itens = carregar_itens_snapshot_lote(numero_lote)

    if not itens:
        return "Nenhum item encontrado para este lote.", 404

    conn = sqlite3.connect("status.db")
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    cursor.execute("""
        SELECT data_criacao
        FROM lotes_envio
        WHERE numero_lote = ?
    """, (numero_lote,))
    lote = cursor.fetchone()
    conn.close()

    data_lote = lote["data_criacao"] if lote else ""

    dados_pdf = []
    for item in itens:
        try:
            dados_item = json.loads(item["dados_json"] or "{}")
        except:
            dados_item = {}

        dados_pdf.append({
            "mlb": item["codigo"],
            "titulo": item["titulo"] or "",
            "sku": item["sku"] or "",
            "quantidade": item["quantidade"] or 0,
            "vendas7": dados_item.get("7 DIAS", 0) or 0,
            "vendas15": dados_item.get("15 DIAS", 0) or 0,
            "vendas30": dados_item.get("30 DIAS", 0) or 0,
            "total_signus": dados_item.get("ESTOQUE TOTAL SIGNUS", 0) or 0,
            "a_caminho": dados_item.get("A CAMINHO DO FULL", 0) or 0,
            "full": dados_item.get("ESTOQUE FULL", 0) or 0,
            "vai_ficar": dados_item.get("ESTOQUE QUE VAI FICAR NO FULL", 0) or 0,
            "comentario": item["comentario_mlb"] or ""
        })

    pdf = gerar_pdf_filete(dados_pdf, data_lote=data_lote)

    return send_file(
        pdf,
        as_attachment=True,
        download_name=f"lote_{numero_lote}.pdf",
        mimetype="application/pdf"
    )

@app.route("/lote-envio/<numero_lote>/filete-excel")
def lote_envio_filete_excel(numero_lote):
    itens = carregar_itens_snapshot_lote(numero_lote)

    if not itens:
        return "Nenhum item encontrado para este lote.", 404

    conn = sqlite3.connect("status.db")
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    cursor.execute("""
        SELECT numero_lote, tipo_lote, data_coleta_agendada
        FROM lotes_envio
        WHERE numero_lote = ?
    """, (numero_lote,))
    lote = cursor.fetchone()
    conn.close()

    tipo_lote = lote["tipo_lote"] if lote else "Diversos"
    data_lote = lote["data_coleta_agendada"] if lote and "data_coleta_agendada" in lote.keys() else ""

    df = pd.DataFrame([{
        "Nickname": item["nickname"],
        "Código do Anúncio": item["codigo"],
        "SKU": item["sku"],
        "Título": item["titulo"],
        "ENDEREÇO": item["endereco"],
        "Enviar": item["quantidade"],
        "Lote": item["lote_filete"]
    } for item in itens])

    output = montar_excel_filete_lote(df, numero_lote, tipo_lote, data_lote)

    return send_file(
        output,
        as_attachment=True,
        download_name=f"filete_{numero_lote}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )



@app.route("/gerar-filete-enviando", methods=["POST"])
def gerar_filete_enviando():
    data = request.get_json() or {}
    numero_lote = str(data.get("numero_lote", "") or "").strip()
    tipo_lote = str(data.get("tipo_lote", "Diversos") or "Diversos").strip() or "Diversos"
    data_lote = str(data.get("data_lote", "") or "").strip()

    if not numero_lote:
        return jsonify({"ok": False, "erro": "Informe o número do lote."}), 400
    if not data_lote:
        return jsonify({"ok": False, "erro": "Informe a data do lote."}), 400

    df = carregar_dados_base()

    conn = sqlite3.connect("status.db")
    cursor = conn.cursor()
    cursor.execute("SELECT codigo, status, quantidade, estrategia FROM status_cards")
    rows = cursor.fetchall()
    conn.close()

    mapa_status = {str(codigo): status or "" for codigo, status, quantidade, estrategia in rows}
    mapa_quantidade = {str(codigo): quantidade or 0 for codigo, status, quantidade, estrategia in rows}

    df["STATUS_CARD"] = df["Código do Anúncio"].astype(str).map(mapa_status).fillna("principal")
    df["Enviar"] = df["Código do Anúncio"].astype(str).map(mapa_quantidade).fillna(0)
    df = df[(df["STATUS_CARD"] == "enviando") & (df["Enviar"].astype(float) > 0)].copy()

    if df.empty:
        return jsonify({"ok": False, "erro": "Não há itens na tela Enviando para gerar o filete."}), 400

    for col in ["Nickname", "Código do Anúncio", "SKU", "Título", "ENDEREÇO"]:
        if col not in df.columns:
            df[col] = ""

    output = montar_excel_filete_antigo(df, numero_lote, tipo_lote, data_lote)

    return send_file(
        output,
        as_attachment=True,
        download_name=f"filete_{numero_lote}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

@app.route("/lote-envio/<numero_lote>/exportar-excel")
def lote_envio_exportar_excel(numero_lote):
    itens = carregar_itens_snapshot_lote(numero_lote)

    if not itens:
        return "Nenhum item encontrado para este lote.", 404

    conn = sqlite3.connect("status.db")
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    cursor.execute("SELECT data_coleta_agendada FROM lotes_envio WHERE numero_lote = ?", (numero_lote,))
    lote = cursor.fetchone()
    conn.close()

    data_coleta_lote = ""
    if lote:
        data_coleta_lote = str(lote["data_coleta_agendada"] or "").strip()

    registros = []
    for item in itens:
        dados_item = json.loads(item["dados_json"] or "{}")
        dados_item["Quantidade para Enviar"] = item["quantidade"] or 0
        dados_item["Estratégia"] = item["estrategia"] or ""
        dados_item["Motivo do Envio"] = item["motivo_envio"] or ""
        dados_item["NUMERO DO LOTE"] = numero_lote
        dados_item["DATA DA COLETA"] = data_coleta_lote
        registros.append(dados_item)

    df = pd.DataFrame(registros)
    df_export = preparar_dataframe_exportacao(df, numero_lote_fixo=numero_lote, data_coleta_fixa=data_coleta_lote)

    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df_export.to_excel(writer, index=False, sheet_name="Relatorio")

    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name=f"relatorio_lote_{numero_lote}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )



@app.route("/dados-full")
def dados_full():
    return jsonify(carregar_csv_com_cache(CSV_URL_FULL, "full"))


@app.route("/salvar-status", methods=["POST"])
def salvar_status():
    data = request.json
    codigo = data.get("codigo")

    conn = sqlite3.connect("status.db")
    cursor = conn.cursor()

    cursor.execute("""
        SELECT status, quantidade, estrategia, motivo_envio, prioridade
        FROM status_cards
        WHERE codigo = ?
    """, (codigo,))
    existente = cursor.fetchone()

    status_atual = "principal"
    quantidade_atual = 0
    estrategia_atual = ""
    motivo_atual = ""
    prioridade_atual = ""

    if existente:
        status_atual = existente[0] or "principal"
        quantidade_atual = existente[1] or 0
        estrategia_atual = existente[2] or ""
        motivo_atual = existente[3] or ""
        prioridade_atual = existente[4] or ""

    status = data.get("status", status_atual)
    quantidade = data.get("quantidade", quantidade_atual)
    estrategia = data.get("estrategia", estrategia_atual)
    motivo_envio = data.get("motivo_envio", motivo_atual)
    prioridade = data.get("prioridade", prioridade_atual)
    if (not str(motivo_envio).strip()) and str(estrategia).strip():
        motivo_envio = estrategia

    try:
        quantidade = int(quantidade)
    except:
        quantidade = 0

    cursor.execute("""
        INSERT INTO status_cards (codigo, status, quantidade, estrategia, motivo_envio, prioridade)
        VALUES (?, ?, ?, ?, ?, ?)
        ON CONFLICT(codigo) DO UPDATE SET
            status=excluded.status,
            quantidade=excluded.quantidade,
            estrategia=excluded.estrategia,
            motivo_envio=excluded.motivo_envio,
            prioridade=excluded.prioridade
    """, (codigo, status, quantidade, estrategia, motivo_envio, prioridade))

    conn.commit()
    conn.close()

    return jsonify({"success": True})


@app.route("/status")
def get_status():
    try:
        conn = sqlite3.connect("status.db")
        cursor = conn.cursor()

        cursor.execute("SELECT codigo, status, quantidade, estrategia, motivo_envio, prioridade FROM status_cards")
        rows = cursor.fetchall()
        mapa_ultimo_lote = obter_ultimo_lote_por_codigo(cursor)

        conn.close()

        status_dict = {}
        for codigo, status, quantidade, estrategia, motivo_envio, prioridade in rows:
            status_dict[str(codigo)] = {
                "status": status,
                "quantidade": quantidade or 0,
                "estrategia": estrategia or "",
                "motivo_envio": motivo_envio or "",
                "prioridade": prioridade or "",
                "ultimo_lote": mapa_ultimo_lote.get(str(codigo), {
                    "numero_lote": "",
                    "quantidade": 0,
                    "data_geracao": "",
                    "status_lote": ""
                })
            }

        for codigo, ultimo_lote in mapa_ultimo_lote.items():
            if codigo not in status_dict:
                status_dict[codigo] = {
                    "status": "principal",
                    "quantidade": 0,
                    "estrategia": "",
                    "motivo_envio": "",
                    "prioridade": "",
                    "ultimo_lote": ultimo_lote
                }

        return jsonify(status_dict)
    except Exception:
        return jsonify({})


@app.route("/comentarios")
def get_comentarios():
    try:
        conn = sqlite3.connect("status.db")
        cursor = conn.cursor()

        cursor.execute("SELECT sku, comentario FROM comentarios")
        rows = cursor.fetchall()

        conn.close()

        comentarios = {str(sku): comentario for sku, comentario in rows}
        return jsonify(comentarios)
    except Exception:
        return jsonify({})


@app.route("/comentarios-mlb")
def get_comentarios_mlb():
    try:
        conn = sqlite3.connect("status.db")
        cursor = conn.cursor()

        cursor.execute("SELECT codigo, comentario FROM comentarios_mlb")
        rows = cursor.fetchall()

        conn.close()

        comentarios = {str(codigo): comentario for codigo, comentario in rows}
        return jsonify(comentarios)
    except Exception:
        return jsonify({})


@app.route("/comentarios-mlb-chat")
def get_comentarios_mlb_chat():
    codigo = str(request.args.get("codigo", "") or "").strip()

    if not codigo:
        return jsonify([])

    try:
        conn = sqlite3.connect("status.db")
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()

        historico = montar_historico_comentarios_mlb(cursor, codigo)

        conn.commit()
        conn.close()

        return jsonify(historico)
    except Exception:
        return jsonify([])


@app.route("/salvar-comentario", methods=["POST"])
def salvar_comentario():
    data = request.json
    sku = data.get("sku")
    comentario = data.get("comentario", "")

    conn = sqlite3.connect("status.db")
    cursor = conn.cursor()

    cursor.execute("""
        INSERT INTO comentarios (sku, comentario)
        VALUES (?, ?)
        ON CONFLICT(sku) DO UPDATE SET comentario=excluded.comentario
    """, (sku, comentario))

    conn.commit()
    conn.close()

    return jsonify({"success": True})


@app.route("/salvar-comentario-mlb", methods=["POST"])
def salvar_comentario_mlb():
    data = request.json or {}
    codigo = str(data.get("codigo", "") or "").strip()
    comentario = str(data.get("comentario", "") or "").strip()

    if not codigo or not comentario:
        return jsonify({"success": False, "erro": "Código e comentário são obrigatórios."}), 400

    conn = sqlite3.connect("status.db")
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()

    historico_existente = montar_historico_comentarios_mlb(cursor, codigo)
    ultimo_texto = ""
    if historico_existente:
        ultimo_texto = str(historico_existente[-1].get("mensagem", "") or "").strip()

    data_hora = agora_str_brasilia()

    if comentario != ultimo_texto:
        cursor.execute("""
            INSERT INTO comentarios_mlb_chat (codigo, mensagem, data_hora)
            VALUES (?, ?, ?)
        """, (codigo, comentario, data_hora))

    cursor.execute("""
        INSERT INTO comentarios_mlb (codigo, comentario)
        VALUES (?, ?)
        ON CONFLICT(codigo) DO UPDATE SET comentario=excluded.comentario
    """, (codigo, comentario))

    conn.commit()
    conn.close()

    return jsonify({
        "success": True,
        "data_hora": data_hora,
        "data_hora_br": formatar_data_hora_br(data_hora, "")
    })


@app.route("/salvar-lote-envio", methods=["POST"])
def salvar_lote_envio():
    data = request.json if request.is_json else request.form

    numero_lote = str(data.get("numero_lote", "") or "").strip()
    tipo_lote = str(data.get("tipo_lote", "") or "").strip() or "Diversos"
    total_mlbs = int(float(data.get("total_mlbs", 0) or 0))
    total_pecas = int(float(data.get("total_pecas", 0) or 0))
    observacao = str(data.get("observacao", "") or "").strip()
    origem = str(data.get("origem", "MANUAL") or "MANUAL").strip() or "MANUAL"
    etapa_atual = str(data.get("etapa_atual", "MLBS EM ANÁLISE") or "MLBS EM ANÁLISE").strip()
    data_coleta_agendada = str(data.get("data_coleta_agendada", "") or "").strip()

    if not numero_lote:
        return jsonify({"ok": False, "erro": "Número do lote é obrigatório."}), 400

    conn = sqlite3.connect("status.db")
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()

    cursor.execute("SELECT * FROM lotes_envio WHERE numero_lote = ?", (numero_lote,))
    existe = cursor.fetchone()

    agora = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    if existe:
        timeline = carregar_timeline_json(existe["timeline_json"])
        etapa_salva = (existe["etapa_atual"] or TIMELINE_ETAPAS[0]).strip() or TIMELINE_ETAPAS[0]
        if not any(str(v or "").strip() for v in timeline.values()):
            timeline[etapa_salva] = existe["data_criacao"] or agora
        timeline = sincronizar_timeline_ate_etapa(timeline, etapa_atual, agora)

        cursor.execute("""
            UPDATE lotes_envio
            SET tipo_lote = ?,
                total_mlbs = ?,
                total_pecas = ?,
                observacao = ?,
                origem = ?,
                etapa_atual = ?,
                timeline_json = ?,
                data_coleta_agendada = ?
            WHERE numero_lote = ?
        """, (
            tipo_lote,
            total_mlbs,
            total_pecas,
            observacao,
            origem,
            etapa_atual,
            timeline_para_json(timeline),
            data_coleta_agendada,
            numero_lote
        ))
    else:
        timeline = timeline_vazio()
        timeline[etapa_atual] = agora

        cursor.execute("""
            INSERT INTO lotes_envio (
                numero_lote, tipo_lote, total_mlbs, total_pecas,
                status, responsavel, transportadora, observacao,
                prioridade, data_envio, status_expedicao, status_ecommerce,
                origem, data_criacao, timeline_json, etapa_atual, data_coleta_agendada
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            numero_lote,
            tipo_lote,
            total_mlbs,
            total_pecas,
            "CRIADO",
            "",
            "",
            observacao,
            "",
            "",
            "AGUARDANDO",
            "AGUARDANDO",
            origem,
            agora,
            timeline_para_json(timeline),
            etapa_atual,
            data_coleta_agendada
        ))

    atualizar_statuss_por_etapa(cursor, numero_lote, etapa_atual)

    conn.commit()
    conn.close()

    if request.headers.get("X-Requested-With") == "XMLHttpRequest" or request.is_json:
        return jsonify({"ok": True})

    return redirect("/metricas-full")


@app.route("/debug-comentarios")
def debug_comentarios():
    conn = sqlite3.connect("status.db")
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM comentarios")
    dados = cursor.fetchall()
    conn.close()
    return jsonify(dados)


@app.route("/dashboard")
def dashboard():
    return render_template(
        "dashboard.html"
    )


from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_LEFT

def gerar_pdf_filete(dados, data_lote=""):
    buffer = BytesIO()

    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        leftMargin=20,
        rightMargin=20,
        topMargin=20,
        bottomMargin=20
    )

    styles = getSampleStyleSheet()

    estilo_titulo = ParagraphStyle(
        name="TituloCustom",
        parent=styles["Title"],
        fontSize=18,
        leading=22,
        alignment=1
    )

    estilo_label = ParagraphStyle(
        name="LabelCustom",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=9,
        leading=11
    )

    estilo_valor = ParagraphStyle(
        name="ValorCustom",
        parent=styles["Normal"],
        fontSize=9,
        leading=11
    )

    estilo_comentario = ParagraphStyle(
        name="ComentarioCustom",
        parent=styles["Normal"],
        fontSize=9,
        leading=12
    )

    elementos = []

    data_pdf = formatar_data_hora_br(data_lote, vazio="-")

    elementos.append(Paragraph("FILETE DE ENVIO", estilo_titulo))
    elementos.append(Spacer(1, 8))
    elementos.append(Paragraph(f"<b>Gerado em:</b> {data_pdf}", estilo_valor))
    elementos.append(Spacer(1, 12))

    for idx, item in enumerate(dados, start=1):
        mlb = str(item.get("mlb", "") or "")
        sku = str(item.get("sku", "") or "")
        qtd = str(item.get("quantidade", "") or "")
        titulo = str(item.get("titulo", "") or "")
        vendas7 = str(item.get("vendas7", "") or "")
        vendas15 = str(item.get("vendas15", "") or "")
        vendas30 = str(item.get("vendas30", "") or "")
        total_signus = str(item.get("total_signus", "") or "0")
        a_caminho = str(item.get("a_caminho", "") or "0")
        full = str(item.get("full", "") or "0")
        vai_ficar = str(item.get("vai_ficar", "") or "0")
        comentario = str(item.get("comentario", "") or "-")

        linha_item = [
    [Paragraph(f"ITEM {idx} - {titulo}", estilo_label)]
]

        tabela_item = Table(linha_item, colWidths=[555])
        tabela_item.setStyle(TableStyle([
            ("GRID", (0, 0), (-1, -1), 0.7, colors.black),
            ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#d9d9d9")),
            ("LEFTPADDING", (0, 0), (-1, -1), 6),
            ("RIGHTPADDING", (0, 0), (-1, -1), 6),
            ("TOPPADDING", (0, 0), (-1, -1), 6),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ]))

        tabela_info = Table([
            [
                Paragraph(f"<b>MLB:</b> {mlb}", estilo_valor),
                Paragraph(f"<b>SKU:</b> {sku}", estilo_valor),
                Paragraph(f"<b>QTD ENVIADA:</b> {qtd}", estilo_valor),
            ],
            [
                Paragraph(f"<b>7 DIAS:</b> {vendas7}", estilo_valor),
                Paragraph(f"<b>15 DIAS:</b> {vendas15}", estilo_valor),
                Paragraph(f"<b>30 DIAS:</b> {vendas30}", estilo_valor),
            ],
            [
                Paragraph(f"<b>TOTAL SIGNUS:</b> {total_signus}", estilo_valor),
                Paragraph(f"<b>A CAMINHO:</b> {a_caminho}", estilo_valor),
                Paragraph(f"<b>FULL:</b> {full}", estilo_valor),
            ],
            [
                Paragraph(f"<b>VAI FICAR:</b> {vai_ficar}", estilo_valor),
                Paragraph(f"<b>DATA LOTE:</b> {data_pdf}", estilo_valor),
                "",
            ]
        ], colWidths=[185, 185, 185])

        tabela_info.setStyle(TableStyle([
            ("GRID", (0, 0), (-1, -1), 0.7, colors.black),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("LEFTPADDING", (0, 0), (-1, -1), 6),
            ("RIGHTPADDING", (0, 0), (-1, -1), 6),
            ("TOPPADDING", (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ]))

        tabela_comentario_titulo = Table([
            [Paragraph("COMENTÁRIO", estilo_label)]
        ], colWidths=[555])

        tabela_comentario_titulo.setStyle(TableStyle([
            ("GRID", (0, 0), (-1, -1), 0.7, colors.black),
            ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#eeeeee")),
            ("LEFTPADDING", (0, 0), (-1, -1), 6),
            ("RIGHTPADDING", (0, 0), (-1, -1), 6),
            ("TOPPADDING", (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ]))

        tabela_comentario = Table([
            [Paragraph(comentario, estilo_comentario)]
        ], colWidths=[555])

        tabela_comentario.setStyle(TableStyle([
            ("GRID", (0, 0), (-1, -1), 0.7, colors.black),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("LEFTPADDING", (0, 0), (-1, -1), 6),
            ("RIGHTPADDING", (0, 0), (-1, -1), 6),
            ("TOPPADDING", (0, 0), (-1, -1), 6),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
        ]))

        elementos.append(tabela_item)
        elementos.append(tabela_info)
        elementos.append(tabela_comentario_titulo)
        elementos.append(tabela_comentario)
        elementos.append(Spacer(1, 12))

    doc.build(elementos)
    buffer.seek(0)
    return buffer

@app.route("/gerar-pdf-filete", methods=["POST"])
def gerar_pdf_filete_route():
    dados = request.json

    pdf = gerar_pdf_filete(dados)

    return send_file(
        pdf,
        as_attachment=True,
        download_name="filete_envio.pdf",
        mimetype="application/pdf"
    )

@app.route("/api/full-distribuicao")
def api_full_distribuicao():
    url = "https://docs.google.com/spreadsheets/d/1DKdRHI9IEacgOwsEd-bnAN4nU3dA_clULxU1mFa8LmY/export?format=csv&gid=46764324"

    try:
        df = pd.read_csv(url)
        df = df.fillna("")
        df.columns = [str(c).strip() for c in df.columns]

        colunas_desejadas = [
            'Unidades que afetam a métrica "Com tempo de estoque"',
            'Entrada pendente',
            'Em transferência',
            'Devolvidas pelo comprador',
            'Não aptas para venda',
            'Temporariamente não aptas para venda\nEnquanto voltam a estar à venda, não ocuparão espaço no Full.',
            'Para colocar à venda',
            'Para evitar descarte'
        ]

        def limpar_numero(valor):
            if pd.isna(valor):
                return 0

            if isinstance(valor, (int, float)):
                return float(valor)

            texto = str(valor).strip()

            if texto == "":
                return 0

            texto = texto.replace(" ", "")

            if "," in texto and "." in texto:
                texto = texto.replace(".", "").replace(",", ".")
            elif "," in texto:
                texto = texto.replace(",", ".")

            try:
                return float(texto)
            except:
                return 0

        dados = []

        for coluna in colunas_desejadas:
            if coluna in df.columns:
                valor_total = df[coluna].apply(limpar_numero).sum()

                if float(valor_total).is_integer():
                    valor_total = int(valor_total)

                dados.append({
                    "titulo": coluna.replace("\n", " "),
                    "coluna": coluna,
                    "valor": valor_total
                })

        return jsonify({
            "ok": True,
            "dados": dados
        })

    except Exception as e:
        return jsonify({
            "ok": False,
            "erro": f"Erro ao carregar distribuição do Full: {str(e)}"
        }), 500

@app.route("/excluir-lote/<numero_lote>", methods=["GET", "POST"])
def excluir_lote(numero_lote):
    conn = sqlite3.connect("status.db")
    cursor = conn.cursor()

    cursor.execute("DELETE FROM lotes_envio WHERE numero_lote = ?", (numero_lote,))
    cursor.execute("DELETE FROM lotes_conferencia WHERE numero_lote = ?", (numero_lote,))
    cursor.execute("DELETE FROM lotes_itens WHERE numero_lote = ?", (numero_lote,))
    cursor.execute("DELETE FROM conferencia_itens WHERE numero_lote = ?", (numero_lote,))
    cursor.execute("DELETE FROM lotes_envio_itens_snapshot WHERE numero_lote = ?", (numero_lote,))

    try:
        cursor.execute("DELETE FROM lotes_picking_itens WHERE numero_lote = ?", (numero_lote,))
    except:
        pass

    conn.commit()
    conn.close()

    return redirect("/metricas-full")

@app.route("/api/full-distribuicao-detalhe")
def api_full_distribuicao_detalhe():
    coluna = str(request.args.get("coluna", "")).strip()

    if not coluna:
        return jsonify({"ok": False, "erro": "Coluna não informada."}), 400

    url_full = "https://docs.google.com/spreadsheets/d/1DKdRHI9IEacgOwsEd-bnAN4nU3dA_clULxU1mFa8LmY/export?format=csv&gid=46764324"

    try:
        df = pd.read_csv(url_full).fillna("")
        df.columns = [str(c).strip() for c in df.columns]

        if coluna not in df.columns:
            return jsonify({"ok": False, "erro": f"Coluna '{coluna}' não encontrada."}), 400

        def encontrar_coluna(df, candidatos):
            colunas_reais = list(df.columns)
            colunas_norm = {str(c).strip().lower(): c for c in colunas_reais}

            for candidato in candidatos:
                chave = str(candidato).strip().lower()
                if chave in colunas_norm:
                    return colunas_norm[chave]

            for col_real in colunas_reais:
                col_real_norm = str(col_real).strip().lower()
                for candidato in candidatos:
                    cand_norm = str(candidato).strip().lower()
                    if cand_norm in col_real_norm:
                        return col_real

            return None

        def limpar_numero(valor):
            if pd.isna(valor):
                return 0

            if isinstance(valor, (int, float)):
                return float(valor)

            texto = str(valor).strip()

            if texto == "":
                return 0

            texto = texto.replace(" ", "")

            if "," in texto and "." in texto:
                texto = texto.replace(".", "").replace(",", ".")
            elif "," in texto:
                texto = texto.replace(",", ".")

            try:
                return float(texto)
            except:
                return 0

        def normalizar_mlb(valor):
            texto = str(valor or "").strip()
            numeros = "".join(ch for ch in texto if ch.isdigit())

            if not numeros:
                return texto

            if texto.upper().startswith("MLB"):
                return f"MLB{numeros}"

            return numeros

        coluna_anuncio = encontrar_coluna(df, [
            "# Anúncio /",
            "# Anúncio",
            "#anúncio",
            "# anuncio",
            "# anúncio /",
            "# anúncio",
            "#anuncio",
            "anúncio",
            "anuncio"
        ])

        coluna_sku = encontrar_coluna(df, [
            "SKU",
            "sku",
            "Sku"
        ])

        coluna_conta = encontrar_coluna(df, [
            "CONTA",
            "Conta",
            "conta"
        ])

        if not coluna_anuncio:
            return jsonify({
                "ok": False,
                "erro": "Coluna '# Anúncio /' não encontrada na aba do Full."
            }), 400

        df["valor_coluna"] = df[coluna].apply(limpar_numero)
        df_filtrado = df[df["valor_coluna"] > 0].copy()

        dados = []
        for _, row in df_filtrado.iterrows():
            unidades = row["valor_coluna"]
            if float(unidades).is_integer():
                unidades = int(unidades)

            mlb = normalizar_mlb(row.get(coluna_anuncio, "")) if coluna_anuncio else ""
            sku = str(row.get(coluna_sku, "")).strip() if coluna_sku else ""
            conta = str(row.get(coluna_conta, "")).strip() if coluna_conta else ""

            dados.append({
                "mlb": mlb,
                "sku": sku,
                "conta": conta,
                "status": coluna.replace("\n", " "),
                "unidades": unidades
            })

        dados = sorted(
            dados,
            key=lambda x: float(x["unidades"]) if x["unidades"] != "" else 0,
            reverse=True
        )

        return jsonify({
            "ok": True,
            "dados": dados
        })

    except Exception as e:
        return jsonify({
            "ok": False,
            "erro": f"Erro ao carregar detalhes da métrica Full: {str(e)}"
        }), 500
    
init_db()

@app.route("/dados-fiscais")
def dados_fiscais():
    """
    Tela de Dados Fiscais por MLB.
    - Sem filtro: mostra todos os MLBs encontrados nos lotes de envio.
    - Com ?lote=NUMERO: mostra somente os MLBs daquele número de lote.
    Os dados fiscais vêm primeiro do dados_json salvo no snapshot do lote e,
    quando possível, são enriquecidos com a base principal.
    """
    numero_lote = str(
        request.args.get("lote")
        or request.args.get("numero_lote")
        or request.args.get("numeroLote")
        or ""
    ).strip()

    resultado = []
    lotes_disponiveis = []

    def pegar(dados, *chaves):
        for chave in chaves:
            if chave in dados and str(dados.get(chave, "") or "").strip() != "":
                return dados.get(chave, "")
        return ""

    try:
        garantir_tabela_lotes_envio_snapshot()

        conn = sqlite3.connect("status.db")
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()

        cursor.execute("""
            SELECT DISTINCT numero_lote
            FROM lotes_envio_itens_snapshot
            WHERE COALESCE(numero_lote, '') <> ''
            ORDER BY numero_lote DESC
        """)
        lotes_disponiveis = [str(row["numero_lote"] or "") for row in cursor.fetchall()]

        if numero_lote:
            cursor.execute("""
                SELECT *
                FROM lotes_envio_itens_snapshot
                WHERE numero_lote = ?
                ORDER BY codigo, sku, id
            """, (numero_lote,))
        else:
            cursor.execute("""
                SELECT *
                FROM lotes_envio_itens_snapshot
                ORDER BY numero_lote DESC, codigo, sku, id
            """)

        itens = [dict(row) for row in cursor.fetchall()]
        conn.close()

        itens = enriquecer_itens_lote_com_base(itens)

        for item in itens:
            try:
                dados_json = json.loads(item.get("dados_json") or "{}")
                if not isinstance(dados_json, dict):
                    dados_json = {}
            except:
                dados_json = {}

            mlb = (
                str(item.get("codigo") or "").strip()
                or str(pegar(dados_json, "MLB", "Código do Anúncio", "# Anúncio", "# de anúncio") or "").strip()
                or "-"
            )

            resultado.append({
                "numero_lote": str(item.get("numero_lote") or ""),
                "mlb": mlb,
                "sku": str(item.get("sku") or pegar(dados_json, "SKU") or ""),
                "titulo": str(item.get("titulo") or pegar(dados_json, "Título", "PRODUTO") or ""),

                "custo_signus": pegar(dados_json, "CUSTO_SIGNUS", "CUSTO SIGNUS", "Custo Signus"),
                "origem_signus": pegar(dados_json, "ORIGEM_SIGNUS", "ORIGEM SIGNUS", "Origem Signus"),
                "ncm": pegar(dados_json, "NCM"),
                "cest_signus": pegar(dados_json, "CEST", "CEST SIGNUS", "CEST_SIGNUS"),
                "st": pegar(dados_json, "COM OU SEM ST", "ST"),
                "id_signus": pegar(dados_json, "ID", "ID SIGNUS", "ID_SIGNUS"),

                "ncm_meli": pegar(dados_json, "NCM MELI", "NCM_MELI"),
                "regra_meli": pegar(dados_json, "REGRA TRIB. MELI", "REGRA TRIB MELI", "REGRA_MELI"),
                "origem_meli": pegar(dados_json, "ORIGEM MELI", "ORIGEM_MELI"),
                "tipo_origem_meli": pegar(dados_json, "TIPO ORIGEM MELI", "TIPO_ORIGEM_MELI"),
                "cest_meli": pegar(dados_json, "CEST MELI", "CEST_MELI"),
                "custo_meli": pegar(dados_json, "CUSTO MELI", "CUSTO_MELI"),
                "id_meli": pegar(dados_json, "ID MELI", "ID_MELI"),
            })

    except Exception as e:
        # Fallback: caso o banco ainda não tenha snapshots, mostra a base FULL.
        try:
            dados_full = carregar_csv_com_cache(CSV_URL_FULL, "full")
        except:
            dados_full = []

        for row in dados_full:
            resultado.append({
                "numero_lote": "",
                "mlb": row.get("MLB") or row.get("Código do Anúncio") or "-",
                "sku": row.get("SKU", ""),
                "titulo": row.get("Título", ""),

                "custo_signus": row.get("CUSTO_SIGNUS", ""),
                "origem_signus": row.get("ORIGEM_SIGNUS", ""),
                "ncm": row.get("NCM", ""),
                "cest_signus": row.get("CEST", "") or row.get("CEST SIGNUS", "") or row.get("CEST_SIGNUS", ""),
                "st": row.get("COM OU SEM ST", ""),
                "id_signus": row.get("ID", ""),

                "ncm_meli": row.get("NCM MELI", ""),
                "regra_meli": row.get("REGRA TRIB. MELI", ""),
                "origem_meli": row.get("ORIGEM MELI", ""),
                "tipo_origem_meli": row.get("TIPO ORIGEM MELI", ""),
                "cest_meli": row.get("CEST MELI", ""),
                "custo_meli": row.get("CUSTO MELI", ""),
                "id_meli": row.get("ID MELI", ""),
            })

    return render_template(
        "dados_fiscais.html",
        dados=resultado,
        lote=numero_lote,
        lotes_disponiveis=lotes_disponiveis,
        total_resultados=len(resultado)
    )


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000, debug=True)